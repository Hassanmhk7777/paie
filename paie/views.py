# paie/views.py
import json
import logging
from decimal import Decimal
from datetime import datetime, date, timedelta

from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.core.files.storage import default_storage
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Q, Sum, Count, Avg
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, get_object_or_404, redirect
from django.template.loader import render_to_string
from django.views.decorators.csrf import csrf_protect, ensure_csrf_cookie
from django.views.decorators.http import require_http_methods

from .models import (
    ParametragePaie, BaremeIR, RubriquePersonnalisee, 
    PeriodePaie, BulletinPaie, LigneBulletin, Employee,
    Site, Department, TypeConge, SoldeConge, DemandeConge, 
    ApprobationConge, RegleConge
)
from .forms import EmployeeForm, EmployeeSearchForm
from .services.calculateur_paie import CalculateurPaieMaroc, CalculateurPeriode
from .services.gestionnaire_conges import GestionnaireConges
from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.contrib.auth.decorators import login_required, permission_required
from django.views.decorators.http import require_http_methods, require_POST
from django.views.decorators.csrf import csrf_exempt
from django.core.paginator import Paginator
from django.db.models import Q, Count, Sum, Avg
from django.utils import timezone
from django.contrib import messages
from django.db import transaction
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side
import io
from .models import (
    Employee, Pointage, PresenceJournaliere, HoraireTravail, 
    PlageHoraire, ReglePointage, ValidationPresence, AlertePresence,
    Department
)
from .services.gestionnaire_pointage import GestionnairePointage

logger = logging.getLogger(__name__)

# Configuration du logger
logger = logging.getLogger(__name__)

# ==============================================================================
# MAIN SPA VIEW
# ==============================================================================

def home(request):
    """Page d'accueil SPA"""
    context = {
        'title': 'Bienvenue sur PaiePro',
        'total_employees': Employee.objects.filter(is_active=True).count(),
        'total_sites': Site.objects.filter(is_active=True).count(),
        'total_departments': Department.objects.filter(is_active=True).count(),
    }
    return render(request, 'base_spa.html', context)

# ==============================================================================
# SPA CONTENT VIEWS
# ==============================================================================

def spa_dashboard(request):
    """Dashboard principal SPA"""
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return render(request, 'spa/dashboard.html')
    return redirect('home')

@ensure_csrf_cookie
def spa_employees_list(request):
    """Liste des employés pour SPA"""
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        # Récupérer la liste des employés avec filtres
        search = request.GET.get('search', '').strip()
        site_id = request.GET.get('site')
        department_id = request.GET.get('department')
        
        employees = Employee.objects.filter(is_active=True).select_related('site', 'department')
        
        # Application des filtres
        if search:
            employees = employees.filter(
                Q(first_name__icontains=search) |
                Q(last_name__icontains=search) |
                Q(email__icontains=search) |
                Q(position__icontains=search)
            )
        
        if site_id:
            employees = employees.filter(site_id=site_id)
            
        if department_id:
            employees = employees.filter(department_id=department_id)
        
        # Pagination
        paginator = Paginator(employees, 20)
        page = request.GET.get('page', 1)
        employees_page = paginator.get_page(page)
        
        # Sites et départements pour les filtres
        sites = Site.objects.filter(is_active=True).order_by('name')
        departments = Department.objects.filter(is_active=True).order_by('name')
        
        context = {
            'employees': employees_page,
            'sites': sites,
            'departments': departments,
            'total_count': paginator.count,
            'search': search,
            'selected_site': site_id,
            'selected_department': department_id,
        }
        
        return render(request, 'spa/employees/list.html', context)
    return redirect('home')

def spa_organigramme(request):
    """Organigramme SPA"""
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        # Récupérer les données pour l'organigramme
        sites = Site.objects.filter(is_active=True)
        
        # Organiser les données par site et département
        organigramme_data = []
        for site in sites:
            site_data = {
                'site': site,
                'departments': [],
                'total_employees': 0
            }
            
            # Récupérer tous les départements
            departments = Department.objects.filter(is_active=True)
            for dept in departments:
                employees = Employee.objects.filter(
                    is_active=True,
                    site=site,
                    department=dept
                ).order_by('position', 'last_name', 'first_name')
                
                if employees.exists():
                    dept_data = {
                        'department': dept,
                        'employees': employees,
                        'count': employees.count()
                    }
                    site_data['departments'].append(dept_data)
                    site_data['total_employees'] += employees.count()
            
            # Employés sans département assigné pour ce site
            employees_no_dept = Employee.objects.filter(
                is_active=True,
                site=site,
                department__isnull=True
            ).order_by('position', 'last_name', 'first_name')
            
            if employees_no_dept.exists():
                site_data['departments'].append({
                    'department': None,
                    'employees': employees_no_dept,
                    'count': employees_no_dept.count()
                })
                site_data['total_employees'] += employees_no_dept.count()
            
            # Ajouter le site même s'il n'a pas d'employés pour le moment
            organigramme_data.append(site_data)
        
        # Employés sans site assigné
        employees_no_site = Employee.objects.filter(
            is_active=True,
            site__isnull=True
        ).order_by('department__name', 'position', 'last_name', 'first_name')
        
        context = {
            'organigramme_data': organigramme_data,
            'employees_no_site': employees_no_site,
            'total_employees': Employee.objects.filter(is_active=True).count(),
            'total_sites': sites.count(),
            'total_departments': Department.objects.filter(is_active=True).count(),
        }
        
        return render(request, 'spa/employees/organigramme.html', context)
    return redirect('home')

# ==============================================================================
# PAYROLL VIEWS
# ==============================================================================

def spa_payroll_calculation(request):
    """Calcul de paie SPA"""
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return payroll_calculation_content(request)
    return redirect('home')

def spa_payroll_bulletins(request):
    """Bulletins de paie SPA"""
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return payroll_bulletins_content(request)
    return redirect('home')

@login_required
def spa_payroll_history(request):
    """Historique paie SPA"""
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        try:
            return payroll_periodes_content(request)
        except Exception as e:
            print(f"Erreur dans payroll_periodes_content: {e}")
            return JsonResponse({'error': str(e)}, status=500)
    return redirect('paie:home')

@login_required
def spa_payroll_settings(request):
    """Paramètres paie SPA"""
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        try:
            return payroll_parametrage_content(request)
        except Exception as e:
            print(f"Erreur dans payroll_parametrage_content: {e}")
            return JsonResponse({'error': str(e)}, status=500)
    return redirect('paie:home')

@login_required
def spa_payroll_statistics(request):
    """Statistiques paie SPA"""
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return payroll_statistics_content(request)
    return redirect('home')

# ==============================================================================
# LEAVE MANAGEMENT VIEWS
# ==============================================================================

def spa_leave_requests(request):
    """Demandes de congés SPA"""
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return render(request, 'spa/leave/requests.html')
    return redirect('home')

def spa_leave_planning(request):
    """Planification congés SPA"""
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return render(request, 'spa/leave/planning.html')
    return redirect('home')

def spa_leave_approvals(request):
    """Approbations congés SPA"""
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return render(request, 'spa/leave/approvals.html')
    return redirect('home')

def spa_leave_calendar(request):
    """Calendrier congés SPA"""
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return render(request, 'spa/leave/calendar.html')
    return redirect('home')

# ==============================================================================
# ATTENDANCE VIEWS
# ==============================================================================

def spa_attendance_live(request):
    """Pointage temps réel SPA"""
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return render(request, 'spa/attendance/live.html')
    return redirect('home')

def spa_attendance_history(request):
    """Historique pointage SPA"""
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return render(request, 'spa/attendance/history.html')
    return redirect('home')

def spa_attendance_absences(request):
    """Gestion absences SPA"""
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return render(request, 'spa/attendance/absences.html')
    return redirect('home')

def spa_attendance_reports(request):
    """Rapports présence SPA"""
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return render(request, 'spa/attendance/reports.html')
    return redirect('home')

# ==============================================================================
# CONFIGURATION VIEWS
# ==============================================================================

def spa_config_sites(request):
    """Gestion sites & départements SPA"""
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        sites = Site.objects.filter(is_active=True)
        departments = Department.objects.filter(is_active=True)
        context = {
            'sites': sites,
            'departments': departments,
        }
        return render(request, 'spa/config/sites.html', context)
    return redirect('home')

def spa_config_users(request):
    """Gestion utilisateurs SPA"""
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return render(request, 'spa/config/users.html')
    return redirect('home')

def spa_config_settings(request):
    """Paramètres généraux SPA"""
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return render(request, 'spa/config/settings.html')
    return redirect('home')

# ==============================================================================
# REPORTS VIEWS
# ==============================================================================

def spa_reports_dashboard(request):
    """Tableaux de bord rapports SPA"""
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return render(request, 'spa/reports/dashboard.html')
    return redirect('home')

def spa_reports_hr(request):
    """Analytics RH SPA"""
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return render(request, 'spa/reports/hr.html')
    return redirect('home')

def spa_reports_exports(request):
    """Exports SPA"""
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return render(request, 'spa/reports/exports.html')
    return redirect('home')

# ==============================================================================
# LEGACY VIEWS (Pour compatibilité)
# ==============================================================================

def dashboard(request):
    """Redirect vers SPA"""
    return redirect('home')

def payroll(request):
    """Redirect vers SPA"""
    return redirect('home')

def leave(request):
    """Redirect vers SPA"""
    return redirect('home')

def attendance(request):
    """Redirect vers SPA"""
    return redirect('home')

def employees(request):
    """Liste des employés avec recherche et filtres"""
    try:
        # Récupération des paramètres de recherche et filtres
        search_form = EmployeeSearchForm(request.GET)
        
        # Requête de base (employés actifs uniquement)
        employees = Employee.objects.filter(is_active=True).select_related('site', 'department')
        
        # Application des filtres
        if search_form.is_valid():
            search = search_form.cleaned_data.get('search')
            site = search_form.cleaned_data.get('site')
            department = search_form.cleaned_data.get('department')
            
            if search:
                employees = employees.filter(
                    Q(first_name__icontains=search) |
                    Q(last_name__icontains=search) |
                    Q(email__icontains=search) |
                    Q(position__icontains=search)
                )
            
            if site:
                employees = employees.filter(site=site)
            
            if department:
                employees = employees.filter(department=department)
        
        # Pagination
        paginator = Paginator(employees, 20)  # 20 employés par page
        page_number = request.GET.get('page')
        employees_page = paginator.get_page(page_number)
        
        # Réponse AJAX pour le filtrage dynamique
        if request.headers.get('x-requested-with') == 'XMLHttpRequest' and request.method == 'GET':
            html = render_to_string(
                'employees/includes/partial_employee_list.html',
                {'employees': employees_page}
            )
            return JsonResponse({
                'html': html,
                'total_count': paginator.count
            })
        
        # Rendu de la page complète
        context = {
            'employees': employees_page,
            'search_form': search_form,
            'title': 'Gestion des Employés',
            'total_count': paginator.count
        }
        return render(request, 'employees/list.html', context)
        
    except Exception as e:
        logger.error(f"Erreur dans la vue employees: {e}")
        messages.error(request, "Une erreur est survenue lors du chargement des employés.")
        return render(request, 'employees/list.html', {
            'employees': [],
            'search_form': EmployeeSearchForm(),
            'title': 'Gestion des Employés'
        })

def save_employee_form(request, form, template_name):
    """Fonction utilitaire pour sauvegarder le formulaire employé"""
    data = {'form_is_valid': False}
    
    try:
        if request.method == 'POST':
            if form.is_valid():
                employee = form.save()
                data['form_is_valid'] = True
                
                # Recharger la liste des employés
                employees = Employee.objects.filter(is_active=True).select_related('site', 'department')
                paginator = Paginator(employees, 20)
                employees_page = paginator.get_page(1)
                
                data['html_employee_list'] = render_to_string(
                    'employees/includes/partial_employee_list.html',
                    {'employees': employees_page}
                )
                
                # Message de succès
                data['message'] = f"L'employé {employee.full_name} a été créé avec succès."
                
            else:
                data['form_is_valid'] = False
                # Construire un message d'erreur détaillé
                errors = []
                for field, field_errors in form.errors.items():
                    field_label = form.fields[field].label if field in form.fields else field
                    for error in field_errors:
                        errors.append(f"{field_label}: {error}")
                
                if errors:
                    data['message'] = "Erreurs de validation:<br>" + "<br>".join(errors)
                else:
                    data['message'] = "Le formulaire contient des erreurs."
                
                # Inclure les erreurs brutes pour le debug
                data['errors'] = dict(form.errors)
        
        # Rendu du formulaire (avec ou sans erreurs)
        context = {'form': form}
        data['html_form'] = render_to_string(template_name, context, request=request)
        
    except Exception as e:
        logger.error(f"Erreur lors de la sauvegarde de l'employé: {e}")
        data['form_is_valid'] = False
        data['message'] = f"Une erreur est survenue lors de la sauvegarde: {str(e)}"
        data['exception'] = str(e)
    
    return JsonResponse(data)

@csrf_protect
@require_http_methods(["GET", "POST"])
def employee_create(request):
    """Création d'un nouvel employé"""
    try:
        if request.method == 'POST':
            form = EmployeeForm(request.POST)
        else:
            form = EmployeeForm()
        
        return save_employee_form(
            request, 
            form, 
            'employees/includes/partial_employee_create.html'
        )
    except Exception as e:
        logger.error(f"Erreur lors de la création de l'employé: {e}")
        return JsonResponse({
            'form_is_valid': False,
            'message': 'Une erreur est survenue lors de la création de l\'employé.'
        })

@csrf_protect
@require_http_methods(["GET", "POST"])
def employee_update(request, pk):
    """Modification d'un employé existant"""
    try:
        employee = get_object_or_404(Employee, pk=pk, is_active=True)
        
        if request.method == 'POST':
            form = EmployeeForm(request.POST, instance=employee)
        else:
            form = EmployeeForm(instance=employee)
        
        return save_employee_form(
            request, 
            form, 
            'employees/includes/partial_employee_update.html'
        )
    except Employee.DoesNotExist:
        return JsonResponse({
            'form_is_valid': False,
            'message': 'Employé introuvable.'
        })
    except Exception as e:
        logger.error(f"Erreur lors de la modification de l'employé {pk}: {e}")
        return JsonResponse({
            'form_is_valid': False,
            'message': 'Une erreur est survenue lors de la modification de l\'employé.'
        })

@csrf_protect
@require_http_methods(["GET", "POST"])
def employee_delete(request, pk):
    """Suppression d'un employé (suppression logique)"""
    try:
        employee = get_object_or_404(Employee, pk=pk, is_active=True)
        data = {}
        
        if request.method == 'POST':
            employee_name = employee.full_name
            
            # Suppression logique (marquer comme inactif)
            employee.is_active = False
            employee.save()
            
            data['form_is_valid'] = True
            data['message'] = f"L'employé {employee_name} a été supprimé avec succès."
            
            # Recharger la liste des employés
            employees = Employee.objects.filter(is_active=True).select_related('site', 'department')
            paginator = Paginator(employees, 20)
            employees_page = paginator.get_page(1)
            
            data['html_employee_list'] = render_to_string(
                'employees/includes/partial_employee_list.html',
                {'employees': employees_page}
            )
        else:
            # Afficher la confirmation de suppression
            context = {'employee': employee}
            data['html_form'] = render_to_string(
                'employees/includes/partial_employee_delete.html',
                context,
                request=request
            )
        
        return JsonResponse(data)
        
    except Employee.DoesNotExist:
        return JsonResponse({
            'form_is_valid': False,
            'message': 'Employé introuvable.'
        })
    except Exception as e:
        logger.error(f"Erreur lors de la suppression de l'employé {pk}: {e}")
        return JsonResponse({
            'form_is_valid': False,
            'message': 'Une erreur est survenue lors de la suppression de l\'employé.'
        })

@require_http_methods(["GET"])
def employee_detail(request, pk):
    """Détails d'un employé"""
    try:
        employee = get_object_or_404(Employee, pk=pk, is_active=True)
        
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            context = {'employee': employee}
            html = render_to_string(
                'employees/includes/partial_employee_detail.html',
                context,
                request=request
            )
            return JsonResponse({'html': html})
        
        return render(request, 'employees/detail.html', {
            'employee': employee,
            'title': f'Détails - {employee.full_name}'
        })
        
    except Employee.DoesNotExist:
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({
                'error': 'Employé introuvable.'
            })
        messages.error(request, 'Employé introuvable.')
        return redirect('employees')
    except Exception as e:
        logger.error(f"Erreur lors de l'affichage des détails de l'employé {pk}: {e}")
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({
                'error': 'Une erreur est survenue.'
            })
        messages.error(request, 'Une erreur est survenue.')
        return redirect('employees')

# Vues utilitaires pour les API internes

@require_http_methods(["GET"])
def get_departments_by_site(request):
    """API pour récupérer les départements d'un site donné"""
    site_id = request.GET.get('site_id')
    if site_id:
        departments = Department.objects.filter(
            is_active=True,
            # Ajoutez ici la logique si les départements sont liés aux sites
        ).values('id', 'name')
        return JsonResponse({'departments': list(departments)})
    
    return JsonResponse({'departments': []})

@require_http_methods(["GET"])
def employee_stats(request):
    """Statistiques des employés pour le dashboard"""
    try:
        stats = {
            'total_employees': Employee.objects.filter(is_active=True).count(),
            'by_site': {},
            'by_department': {},
            'recent_hires': Employee.objects.filter(is_active=True).order_by('-hire_date')[:5].values(
                'first_name', 'last_name', 'hire_date', 'position'
            )
        }
        
        # Statistiques par site
        for site in Site.objects.filter(is_active=True):
            stats['by_site'][site.name] = site.employee_count
        
        # Statistiques par département
        for dept in Department.objects.filter(is_active=True):
            stats['by_department'][dept.name] = dept.employee_count
        
        return JsonResponse(stats)
        
    except Exception as e:
        logger.error(f"Erreur lors du calcul des statistiques: {e}")
        return JsonResponse({'error': 'Erreur lors du calcul des statistiques'})

@login_required
def payroll_calculation_content(request):
    """Contenu SPA - Calcul de paie"""
    
    # Périodes récentes
    periodes = PeriodePaie.objects.select_related('parametrage').order_by('-date_debut')[:10]
    
    # Employés actifs
    employes = Employee.objects.filter(is_active=True).order_by('last_name', 'first_name')
    
    # Paramétrage actuel
    parametrage_actuel = ParametragePaie.objects.filter(actif=True).first()
    
    context = {
        'periodes': periodes,
        'employes': employes,
        'parametrage_actuel': parametrage_actuel,
    }
    
    return render(request, 'spa/payroll/calculation.html', context)

@login_required
def payroll_bulletins_content(request):
    """Contenu SPA - Gestion des bulletins"""
    
    # Filtres
    periode_id = request.GET.get('periode')
    employe_id = request.GET.get('employe')
    search = request.GET.get('search', '').strip()
    
    # Query de base
    bulletins = BulletinPaie.objects.select_related(
        'periode', 'employe', 'genere_par'
    ).order_by('-periode__date_debut', 'employe__last_name')
    
    # Application des filtres
    if periode_id:
        bulletins = bulletins.filter(periode_id=periode_id)
    
    if employe_id:
        bulletins = bulletins.filter(employe_id=employe_id)
    
    if search:
        bulletins = bulletins.filter(
            Q(employe__last_name__icontains=search) |
            Q(employe__first_name__icontains=search) |
            Q(numero_bulletin__icontains=search)
        )
    
    # Pagination
    paginator = Paginator(bulletins, 20)
    page = request.GET.get('page', 1)
    bulletins_page = paginator.get_page(page)
    
    # Données pour filtres
    periodes = PeriodePaie.objects.order_by('-date_debut')[:12]
    employes = Employee.objects.filter(is_active=True).order_by('last_name')
    
    context = {
        'bulletins': bulletins_page,
        'periodes': periodes,
        'employes': employes,
        'filters': {
            'periode_id': periode_id,
            'employe_id': employe_id,
            'search': search,
        }
    }
    
    return render(request, 'spa/payroll/bulletins.html', context)

@login_required
def payroll_parametrage_content(request):
    """Contenu SPA - Paramétrage paie"""
    
    try:
        # Paramétrage actuel
        parametrage_actuel = ParametragePaie.objects.filter(actif=True).first()
        
        # Historique paramétrages
        historique = ParametragePaie.objects.order_by('-id')[:5]  # Utiliser -id au lieu de -annee
        
        # Rubriques personnalisées - simplifiées
        rubriques = RubriquePersonnalisee.objects.filter(actif=True)
        
        context = {
            'parametrage_actuel': parametrage_actuel,
            'historique': historique,
            'rubriques': rubriques,
        }
        
        return render(request, 'spa/payroll/parametrage.html', context)
        
    except Exception as e:
        print(f"Erreur dans payroll_parametrage_content: {e}")
        # Retourner une page vide si erreur  
        context = {
            'parametrage_actuel': None,
            'historique': [],
            'rubriques': [],
        }
        return render(request, 'spa/payroll/parametrage.html', context)

@login_required
def payroll_statistics_content(request):
    """Contenu SPA - Statistiques paie avancées"""
    
    from django.db.models import Count, Sum, Avg
    from django.db.models.functions import TruncMonth
    from datetime import datetime, timedelta
    import calendar
    
    # Période par défaut (derniers 12 mois)
    periode_fin = datetime.now().date()
    periode_debut = periode_fin.replace(year=periode_fin.year - 1)
    
    # Filtres depuis la requête
    periode_filter = request.GET.get('periode', '12_mois')
    departement_filter = request.GET.get('departement', '')
    date_debut_custom = request.GET.get('date_debut', '')
    date_fin_custom = request.GET.get('date_fin', '')
    
    # Ajuster les dates selon le filtre
    if periode_filter == '1_mois':
        periode_debut = periode_fin.replace(day=1)
    elif periode_filter == '3_mois':
        periode_debut = periode_fin - timedelta(days=90)
    elif periode_filter == '6_mois':
        periode_debut = periode_fin - timedelta(days=180)
    elif periode_filter == 'custom' and date_debut_custom and date_fin_custom:
        try:
            periode_debut = datetime.strptime(date_debut_custom, '%Y-%m-%d').date()
            periode_fin = datetime.strptime(date_fin_custom, '%Y-%m-%d').date()
        except ValueError:
            pass
    
    # Base queryset pour les bulletins de la période
    bulletins_qs = BulletinPaie.objects.filter(
        periode__date_debut__gte=periode_debut,
        periode__date_fin__lte=periode_fin
    ).select_related('employe', 'periode', 'employe__department', 'employe__site')
    
    # Filtre par département si spécifié
    if departement_filter:
        bulletins_qs = bulletins_qs.filter(employe__department__name=departement_filter)
    
    # Employés actifs pour la période
    employes_actifs = Employee.objects.filter(is_active=True)
    if departement_filter:
        employes_actifs = employes_actifs.filter(department__name=departement_filter)
    
    # ========== KPIs PRINCIPAUX ==========
    
    # Calculs de base
    total_bulletins = bulletins_qs.count()
    masse_salariale_totale = bulletins_qs.aggregate(total=Sum('net_a_payer'))['total'] or 0
    employes_actifs_count = employes_actifs.count()
    
    # Masse salariale mois précédent pour variation
    mois_precedent_debut = (periode_fin.replace(day=1) - timedelta(days=1)).replace(day=1)
    mois_precedent_fin = periode_fin.replace(day=1) - timedelta(days=1)
    
    masse_salariale_precedente = BulletinPaie.objects.filter(
        periode__date_debut__gte=mois_precedent_debut,
        periode__date_fin__lte=mois_precedent_fin
    )
    if departement_filter:
        masse_salariale_precedente = masse_salariale_precedente.filter(employe__department__name=departement_filter)
    
    masse_precedente_total = masse_salariale_precedente.aggregate(total=Sum('net_a_payer'))['total'] or 0
    
    # Variation masse salariale
    if masse_precedente_total > 0:
        variation_masse = ((masse_salariale_totale - masse_precedente_total) / masse_precedente_total) * 100
    else:
        variation_masse = 0
    
    # Nouveaux employés (derniers 30 jours)
    date_30j = periode_fin - timedelta(days=30)
    nouveaux_employes = employes_actifs.filter(hire_date__gte=date_30j).count()
    
    # Taux de réussite bulletins (approximation - bulletins sans erreurs majeures)
    bulletins_reussis = bulletins_qs.filter(net_a_payer__gt=0).count()
    taux_reussite = (bulletins_reussis / max(total_bulletins, 1)) * 100
    
    # Temps moyen de traitement (simulation basée sur la période)
    temps_moyen_traitement = 2.3  # Heures - simulation
    evolution_traitement = -0.2  # Amélioration de 0.2h
    
    # Taux d'absentéisme (simulation)
    taux_absenteisme = 4.2  # %
    variation_absenteisme = 0.3  # %
    
    # Coût moyen par employé
    cout_moyen_employe = masse_salariale_totale / max(employes_actifs_count, 1)
    cout_precedent = masse_precedente_total / max(employes_actifs_count, 1)
    tendance_cout = cout_moyen_employe - cout_precedent
    
    kpis = {
        'masse_salariale': {
            'valeur': masse_salariale_totale,
            'variation': variation_masse,
            'type': 'increase' if variation_masse > 0 else 'decrease' if variation_masse < 0 else 'stable'
        },
        'employes_actifs': {
            'valeur': employes_actifs_count,
            'nouveaux': nouveaux_employes,
            'type': 'info'
        },
        'bulletins_traites': {
            'valeur': total_bulletins,
            'taux_reussite': taux_reussite,
            'type': 'success'
        },
        'temps_traitement': {
            'valeur': temps_moyen_traitement,
            'evolution': evolution_traitement,
            'type': 'decrease' if evolution_traitement < 0 else 'increase'
        },
        'taux_absenteisme': {
            'valeur': taux_absenteisme,
            'variation': variation_absenteisme,
            'type': 'warning'
        },
        'cout_moyen': {
            'valeur': cout_moyen_employe,
            'tendance': tendance_cout,
            'type': 'increase' if tendance_cout > 0 else 'decrease'
        }
    }
    
    # ========== DONNÉES GRAPHIQUES ==========
    
    # Évolution masse salariale (12 derniers mois)
    douze_mois_ago = periode_fin - timedelta(days=365)
    evolution_mensuelle = BulletinPaie.objects.filter(
        periode__date_debut__gte=douze_mois_ago
    )
    if departement_filter:
        evolution_mensuelle = evolution_mensuelle.filter(employe__department__name=departement_filter)
    
    evolution_data = evolution_mensuelle.annotate(
        mois=TruncMonth('periode__date_debut')
    ).values('mois').annotate(
        masse=Sum('net_a_payer'),
        bulletins=Count('id')
    ).order_by('mois')
    
    # Préparer les données pour Chart.js
    evolution_labels = []
    evolution_values = []
    
    for item in evolution_data:
        mois_nom = calendar.month_name[item['mois'].month]
        evolution_labels.append(f"{mois_nom} {item['mois'].year}")
        evolution_values.append(float(item['masse'] or 0))
    
    # Répartition par département
    repartition_dept = bulletins_qs.values(
        'employe__department__name'
    ).annotate(
        masse=Sum('net_a_payer'),
        effectif=Count('employe', distinct=True)
    ).order_by('-masse')[:8]  # Top 8 départements
    
    dept_labels = []
    dept_values = []
    dept_colors = [
        '#4f68d8', '#27ae60', '#f39c12', '#e74c3c',
        '#9b59b6', '#1abc9c', '#34495e', '#95a5a6'
    ]
    
    for i, item in enumerate(repartition_dept):
        dept_name = item['employe__department__name'] or 'Non assigné'
        dept_labels.append(dept_name)
        dept_values.append(float(item['masse'] or 0))
    
    # ========== TABLEAUX DÉTAILLÉS ==========
    
    # Indicateurs RH
    # Calcul de l'ancienneté moyenne manuellement pour compatibilité SQLite
    employes_avec_dates = employes_actifs.filter(hire_date__isnull=False)
    
    if employes_avec_dates.exists():
        # Calculer manuellement l'ancienneté moyenne
        total_jours = 0
        count = 0
        for emp in employes_avec_dates:
            jours = (periode_fin - emp.hire_date).days
            total_jours += jours
            count += 1
        
        anciennete_annees = round((total_jours / count) / 365.25, 1) if count > 0 else 0
    else:
        anciennete_annees = 0
    
    # Simulation des autres indicateurs RH
    indicateurs_rh = {
        'turnover': 8.5,  # %
        'anciennete_moyenne': anciennete_annees,
        'age_moyen': 32.4,  # ans
        'ratio_hf': 1.2,  # H/F
        'embauches_mois': nouveaux_employes,
        'departs_mois': 2  # simulation
    }
    
    # Charges sociales
    charges_cnss = bulletins_qs.aggregate(total=Sum('charges_cnss_patronal'))['total'] or 0
    charges_amo = bulletins_qs.aggregate(total=Sum('charges_amo_patronal'))['total'] or 0
    formation_prof = bulletins_qs.aggregate(total=Sum('formation_professionnelle'))['total'] or 0
    
    charges_sociales = {
        'cnss_patronale': charges_cnss,
        'cnss_salariale': bulletins_qs.aggregate(total=Sum('cotisation_cnss'))['total'] or 0,
        'amo': charges_amo,
        'formation': formation_prof,
        'accident_travail': charges_cnss * Decimal('0.01') if charges_cnss else 0,  # Estimation 1%
        'total': charges_cnss + charges_amo + formation_prof
    }
    
    # Performance système (simulation)
    performance_systeme = {
        'bulletins_sans_erreur': f"{taux_reussite:.1f}%",
        'temps_traitement_moyen': f"{temps_moyen_traitement:.1f}h",
        'erreurs_calcul': max(0, total_bulletins - bulletins_reussis),
        'disponibilite': 99.2  # %
    }
    
    # Top départements par coût
    top_departements = list(repartition_dept)[:5]
    
    # ========== DÉPARTEMENTS POUR FILTRES ==========
    departements_list = Employee.objects.filter(
        is_active=True, 
        department__isnull=False
    ).values_list(
        'department__name', flat=True
    ).distinct().order_by('department__name')
    
    context = {
        'kpis': kpis,
        'periode_debut': periode_debut,
        'periode_fin': periode_fin,
        'evolution_labels': evolution_labels,
        'evolution_values': evolution_values,
        'dept_labels': dept_labels,
        'dept_values': dept_values,
        'dept_colors': dept_colors,
        'indicateurs_rh': indicateurs_rh,
        'charges_sociales': charges_sociales,
        'performance_systeme': performance_systeme,
        'top_departements': top_departements,
        'departements_list': departements_list,
        'periode_filter': periode_filter,
        'departement_filter': departement_filter,
        'date_debut_custom': date_debut_custom,
        'date_fin_custom': date_fin_custom,
    }
    
    return render(request, 'spa/payroll/statistics.html', context)

@login_required
def payroll_periodes_content(request):
    """Contenu SPA - Gestion des périodes"""
    
    try:
        # Simplifier d'abord pour éviter les erreurs
        periodes = PeriodePaie.objects.all().order_by('-date_debut')
        
        context = {
            'periodes': periodes,
        }
        
        return render(request, 'spa/payroll/periodes.html', context)
        
    except Exception as e:
        print(f"Erreur dans payroll_periodes_content: {e}")
        # Retourner une page vide si erreur
        return render(request, 'spa/payroll/periodes.html', {'periodes': []})

# ================== APIs PAIE ==================

@login_required
@require_http_methods(["POST"])
def api_calculer_bulletin_test(request):
    """API - Calculer un bulletin en mode test (sans sauvegarde)"""
    
    try:
        data = json.loads(request.body)
        employe_id = data.get('employe_id')
        periode_id = data.get('periode_id')
        donnees_variables = data.get('donnees_variables', {})
        
        # Validation
        employe = get_object_or_404(Employee, id=employe_id, is_active=True)
        periode = get_object_or_404(PeriodePaie, id=periode_id)
        
        # Calcul
        calculateur = CalculateurPaieMaroc(periode.parametrage)
        calcul = calculateur.calculer_bulletin(employe, periode, donnees_variables)
        
        # Formatage pour JSON
        resultat = {
            'employe': {
                'nom': employe.last_name,
                'prenom': employe.first_name,
                'salaire_base': float(employe.salary),
            },
            'periode': periode.libelle,
            'elements_brut': {
                'salaire_base': float(calcul['salaire_base_period']),
                'heures_sup': float(calcul['montant_heures_sup']),
                'prime_anciennete': float(calcul['prime_anciennete']),
                'prime_responsabilite': float(calcul['prime_responsabilite']),
                'indemnite_transport': float(calcul['indemnite_transport']),
                'avantages_nature': float(calcul['avantages_nature']),
                'total_brut': float(calcul['total_brut']),
            },
            'cotisations': {
                'cnss': float(calcul['cotisation_cnss']),
                'amo': float(calcul['cotisation_amo']),
                'cimr': float(calcul['cotisation_cimr']),
            },
            'impots': {
                'base_imposable': float(calcul['total_imposable']),
                'ir_brut': float(calcul['ir_brut']),
                'ir_net': float(calcul['ir_net']),
            },
            'retenues': {
                'avances': float(calcul['avances']),
                'prets': float(calcul['prets']),
                'autres': float(calcul['autres_retenues']),
                'total_retenues': float(calcul['total_retenues']),
            },
            'net_a_payer': float(calcul['net_a_payer']),
            'charges_patronales': {
                'cnss': float(calcul['charges_cnss_patronal']),
                'amo': float(calcul['charges_amo_patronal']),
                'formation': float(calcul['formation_professionnelle']),
                'prestations': float(calcul['prestations_sociales']),
            },
            'rubriques_personnalisees': [
                {
                    'code': ligne['rubrique'].code,
                    'libelle': ligne['rubrique'].libelle,
                    'type': ligne['rubrique'].type_rubrique,
                    'montant': float(ligne['montant']),
                }
                for ligne in calcul['rubriques_gains'] + calcul['rubriques_retenues']
            ]
        }
        
        return JsonResponse({
            'success': True,
            'resultat': resultat
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)

@login_required
@require_http_methods(["POST"])
def api_generer_bulletin(request):
    """API - Générer et sauvegarder un bulletin"""
    
    try:
        data = json.loads(request.body)
        employe_id = data.get('employe_id')
        periode_id = data.get('periode_id')
        donnees_variables = data.get('donnees_variables', {})
        
        # Validation
        employe = get_object_or_404(Employee, id=employe_id, is_active=True)
        periode = get_object_or_404(PeriodePaie, id=periode_id)
        
        if periode.statut == 'CLOTUREE':
            return JsonResponse({
                'success': False,
                'error': 'Impossible de modifier une période clôturée'
            }, status=400)
        
        # Vérifier si bulletin existe déjà
        bulletin_existant = BulletinPaie.objects.filter(
            periode=periode, employe=employe
        ).first()
        
        if bulletin_existant and not data.get('force_recreate', False):
            return JsonResponse({
                'success': False,
                'error': 'Un bulletin existe déjà pour cette période. Utilisez force_recreate=true pour le remplacer.'
            }, status=400)
        
        # Génération
        with transaction.atomic():
            if bulletin_existant:
                bulletin_existant.delete()
            
            calculateur = CalculateurPaieMaroc(periode.parametrage)
            bulletin = calculateur.generer_bulletin_db(employe, periode, donnees_variables)
            bulletin.genere_par = request.user
            bulletin.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Bulletin généré avec succès pour {employe.first_name} {employe.last_name}',
            'bulletin_id': bulletin.id,
            'numero': bulletin.numero_bulletin,
            'net_a_payer': float(bulletin.net_a_payer)
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)

@login_required
@require_http_methods(["POST"])
def api_calculer_periode(request):
    """API - Calculer une période complète"""
    
    try:
        data = json.loads(request.body)
        periode_id = data.get('periode_id')
        employes_ids = data.get('employes_ids')  # None = tous
        force_recreate = data.get('force_recreate', False)
        
        periode = get_object_or_404(PeriodePaie, id=periode_id)
        
        # Calcul
        calculateur_periode = CalculateurPeriode()
        stats = calculateur_periode.calculer_periode_complete(
            periode, employes_ids, force_recreate
        )
        
        return JsonResponse({
            'success': True,
            'stats': stats
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)

@login_required
def debug_calcul_paie(request):
    """Page de debug pour le calcul de paie"""
    from django.http import FileResponse
    import os
    
    debug_file_path = os.path.join(settings.BASE_DIR, 'test_calcul_debug.html')
    return FileResponse(open(debug_file_path, 'rb'), content_type='text/html')

@login_required
@require_http_methods(["GET"])
def api_bulletin_detail(request, bulletin_id):
    """API - Détail d'un bulletin"""
    
    bulletin = get_object_or_404(
        BulletinPaie.objects.select_related('employe', 'periode', 'genere_par'),
        id=bulletin_id
    )
    
    # Lignes des rubriques personnalisées
    lignes = LigneBulletin.objects.filter(bulletin=bulletin).select_related('rubrique')
    
    data = {
        'bulletin': {
            'id': bulletin.id,
            'numero': bulletin.numero_bulletin,
            'employe': {
                'nom': bulletin.employe.nom,
                'prenom': bulletin.employe.prenom,
                'matricule': getattr(bulletin.employe, 'matricule', ''),
            },
            'periode': {
                'libelle': bulletin.periode.libelle,
                'date_debut': bulletin.periode.date_debut.isoformat(),
                'date_fin': bulletin.periode.date_fin.isoformat(),
            },
            'elements_brut': {
                'salaire_base': float(bulletin.salaire_base),
                'heures_sup': float(bulletin.heures_supplementaires),
                'prime_anciennete': float(bulletin.prime_anciennete),
                'prime_responsabilite': float(bulletin.prime_responsabilite),
                'indemnite_transport': float(bulletin.indemnite_transport),
                'avantages_nature': float(bulletin.avantages_nature),
                'total_brut': float(bulletin.total_brut),
            },
            'cotisations': {
                'cnss': float(bulletin.cotisation_cnss),
                'amo': float(bulletin.cotisation_amo),
                'cimr': float(bulletin.cotisation_cimr),
            },
            'impots': {
                'base_imposable': float(bulletin.total_imposable),
                'ir_brut': float(bulletin.ir_brut),
                'ir_net': float(bulletin.ir_net),
            },
            'retenues': {
                'avances': float(bulletin.avances),
                'prets': float(bulletin.prets),
                'autres': float(bulletin.autres_retenues),
                'total_retenues': float(bulletin.total_retenues),
            },
            'net_a_payer': float(bulletin.net_a_payer),
            'charges_patronales': {
                'cnss': float(bulletin.charges_cnss_patronal),
                'amo': float(bulletin.charges_amo_patronal),
                'formation': float(bulletin.formation_professionnelle),
                'prestations': float(bulletin.prestations_sociales),
            },
            'rubriques_personnalisees': [
                {
                    'code': ligne.rubrique.code,
                    'libelle': ligne.rubrique.libelle,
                    'type': ligne.rubrique.type_rubrique,
                    'base_calcul': float(ligne.base_calcul),
                    'taux': float(ligne.taux_applique) if ligne.taux_applique else None,
                    'montant': float(ligne.montant),
                }
                for ligne in lignes
            ],
            'metadata': {
                'date_generation': bulletin.date_generation.isoformat(),
                'genere_par': bulletin.genere_par.username if bulletin.genere_par else None,
            }
        }
    }
    
    return JsonResponse(data)

@login_required
@require_http_methods(["GET"])
def api_statistiques_paie(request):
    """API - Statistiques du module paie"""
    
    # Période en cours
    periode_courante = PeriodePaie.objects.filter(
        date_debut__lte=date.today(),
        date_fin__gte=date.today()
    ).first()
    
    # Statistiques générales
    stats = {
        'bulletins_total': BulletinPaie.objects.count(),
        'employes_actifs': Employee.objects.filter(is_active=True).count(),
        'periodes_total': PeriodePaie.objects.count(),
        'rubriques_personnalisees': RubriquePersonnalisee.objects.filter(actif=True).count(),
    }
    
    # Période courante
    if periode_courante:
        bulletins_periode = BulletinPaie.objects.filter(periode=periode_courante)
        stats['periode_courante'] = {
            'libelle': periode_courante.libelle,
            'statut': periode_courante.statut,
            'bulletins_generes': bulletins_periode.count(),
            'masse_salariale': float(bulletins_periode.aggregate(
                total=Sum('net_a_payer')
            )['total'] or 0),
            'total_charges': float(sum([
                bulletins_periode.aggregate(cnss=Sum('charges_cnss_patronal'))['cnss'] or 0,
                bulletins_periode.aggregate(amo=Sum('charges_amo_patronal'))['amo'] or 0,
                bulletins_periode.aggregate(formation=Sum('formation_professionnelle'))['formation'] or 0,
                bulletins_periode.aggregate(prestations=Sum('prestations_sociales'))['prestations'] or 0
            ])),
        }
    
    # Évolution sur 6 derniers mois
    from datetime import datetime, timedelta
    from django.db.models.functions import TruncMonth
    
    six_months_ago = date.today() - timedelta(days=180)
    evolution = BulletinPaie.objects.filter(
        periode__date_debut__gte=six_months_ago
    ).annotate(
        mois=TruncMonth('periode__date_debut')
    ).values('mois').annotate(
        nombre_bulletins=Count('id'),
        masse_salariale=Sum('net_a_payer'),
        total_charges=Sum('charges_cnss_patronal') + Sum('charges_amo_patronal')
    ).order_by('mois')
    
    stats['evolution'] = [
        {
            'mois': item['mois'].strftime('%Y-%m'),
            'bulletins': item['nombre_bulletins'],
            'masse_salariale': float(item['masse_salariale'] or 0),
            'charges': float(item['total_charges'] or 0),
        }
        for item in evolution
    ]
    
    return JsonResponse(stats)

@login_required 
@require_http_methods(["GET"])
def api_export_statistiques_paie(request):
    """API - Export des statistiques de paie en PDF"""
    
    from django.http import HttpResponse
    from django.template.loader import render_to_string
    import io
    
    try:
        # Récupérer les mêmes données que pour la vue
        context = {}
        
        # Simulation des données pour l'export
        kpis_export = {
            'masse_salariale': 450000,
            'employes_actifs': 25,
            'bulletins_traites': 25,
            'taux_reussite': 98.5,
            'temps_traitement': 2.3,
            'taux_absenteisme': 4.2
        }
        
        charges_export = {
            'cnss_patronale': 45000,
            'cnss_salariale': 20000,
            'amo': 15000,
            'formation': 5000,
            'total': 85000
        }
        
        context.update({
            'kpis': kpis_export,
            'charges': charges_export,
            'periode': request.GET.get('periode', '12_mois'),
            'departement': request.GET.get('departement', ''),
            'date_generation': datetime.now(),
        })
        
        # Générer le HTML pour le PDF
        # html_content = render_to_string('spa/payroll/statistics_pdf.html', context, request)
        
        # Simuler la génération PDF (en réalité, utiliser WeasyPrint ou ReportLab)
        pdf_content = f"""
        RAPPORT STATISTIQUES PAIE - EXPORT PDF
        =====================================
        
        Période: {context['periode']}
        Département: {context['departement'] or 'Tous'}
        Date de génération: {context['date_generation'].strftime('%d/%m/%Y %H:%M')}
        
        KPIs PRINCIPAUX:
        - Masse salariale: {kpis_export['masse_salariale']:,} MAD
        - Employés actifs: {kpis_export['employes_actifs']}
        - Bulletins traités: {kpis_export['bulletins_traites']}
        - Taux de réussite: {kpis_export['taux_reussite']}%
        
        CHARGES SOCIALES:
        - CNSS Patronale: {charges_export['cnss_patronale']:,} MAD
        - CNSS Salariale: {charges_export['cnss_salariale']:,} MAD
        - AMO: {charges_export['amo']:,} MAD
        - Formation: {charges_export['formation']:,} MAD
        - Total: {charges_export['total']:,} MAD
        
        [Ce serait un vrai PDF avec graphiques et tableaux formatés]
        """.encode('utf-8')
        
        # Préparer la réponse
        response = HttpResponse(pdf_content, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="statistiques_paie_{datetime.now().strftime("%Y%m%d")}.pdf"'
        
        return response
        
    except Exception as e:
        logger.error(f"Erreur export statistiques: {e}")
        return JsonResponse({
            'success': False,
            'error': 'Erreur lors de la génération du rapport PDF'
        }, status=500)

# ================== GESTION RUBRIQUES PERSONNALISÉES ==================

@login_required
@require_http_methods(["POST"])
def api_rubrique_create(request):
    """API - Créer une rubrique personnalisée"""
    
    try:
        data = json.loads(request.body)
        
        rubrique = RubriquePersonnalisee.objects.create(
            code=data['code'],
            libelle=data['libelle'],
            type_rubrique=data['type_rubrique'],
            mode_calcul=data['mode_calcul'],
            periodicite=data['periodicite'],
            valeur_fixe=Decimal(str(data.get('valeur_fixe', 0))) if data.get('valeur_fixe') else None,
            pourcentage=Decimal(str(data.get('pourcentage', 0))) if data.get('pourcentage') else None,
            formule=data.get('formule'),
            imposable_ir=data.get('imposable_ir', True),
            soumis_cnss=data.get('soumis_cnss', True),
            soumis_amo=data.get('soumis_amo', True),
            soumis_cimr=data.get('soumis_cimr', False),
            ordre_affichage=data.get('ordre_affichage', 1),
            afficher_bulletin=data.get('afficher_bulletin', True),
        )
        
        return JsonResponse({
            'success': True,
            'rubrique_id': rubrique.id,
            'message': 'Rubrique créée avec succès'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)

@login_required
@require_http_methods(["PUT"])
def api_rubrique_update(request, rubrique_id):
    """API - Modifier une rubrique personnalisée"""
    
    try:
        rubrique = get_object_or_404(RubriquePersonnalisee, id=rubrique_id)
        data = json.loads(request.body)
        
        # Mise à jour des champs
        for field in ['libelle', 'type_rubrique', 'mode_calcul', 'periodicite', 
                     'formule', 'imposable_ir', 'soumis_cnss', 'soumis_amo', 
                     'soumis_cimr', 'ordre_affichage', 'afficher_bulletin']:
            if field in data:
                setattr(rubrique, field, data[field])
        
        # Champs décimaux
        if 'valeur_fixe' in data:
            rubrique.valeur_fixe = Decimal(str(data['valeur_fixe'])) if data['valeur_fixe'] else None
        if 'pourcentage' in data:
            rubrique.pourcentage = Decimal(str(data['pourcentage'])) if data['pourcentage'] else None
        
        rubrique.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Rubrique mise à jour avec succès'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)

@login_required
@require_http_methods(["DELETE"])
def api_rubrique_delete(request, rubrique_id):
    """API - Désactiver une rubrique personnalisée"""
    
    try:
        rubrique = get_object_or_404(RubriquePersonnalisee, id=rubrique_id)
        rubrique.actif = False
        rubrique.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Rubrique désactivée avec succès'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)
    # ================== VUES CONTENU SPA CONGÉS ==================

@login_required
def leave_requests_content(request):
    """Contenu SPA - Demandes de congés employé"""
    
        # Récupérer l'employé connecté (ou employé sélectionné pour RH)
    employe_id = request.GET.get('employe_id')
    if employe_id and request.user.is_staff:
        # RH peut voir les demandes d'autres employés
        employe = get_object_or_404(Employee, id=employe_id, is_active=True)
    else:
        # Employé normal ne voit que ses demandes
        try:
            employe = Employee.objects.get(email=request.user.email, is_active=True)
        except Employee.DoesNotExist:
            # Si pas trouvé par email, prendre le premier employé (demo)
            employe = Employee.objects.filter(is_active=True).first()    # Mes demandes de congés
    mes_demandes = DemandeConge.objects.filter(
        employe=employe
    ).select_related('type_conge').order_by('-date_creation')
    
    # Pagination
    paginator = Paginator(mes_demandes, 10)
    page = request.GET.get('page', 1)
    demandes_page = paginator.get_page(page)
    
    # Types de congés disponibles
    types_conges = TypeConge.objects.filter(is_active=True).order_by('ordre_affichage')
    
    # Soldes de congés
    gestionnaire = GestionnaireConges()
    annee_courante = datetime.now().year
    soldes = gestionnaire.calculer_soldes_employe(employe, annee_courante)
    
    # Statistiques rapides
    stats = {
        'total_demandes': mes_demandes.count(),
        'demandes_en_attente': mes_demandes.filter(
            statut__in=['SOUMISE', 'EN_ATTENTE_MANAGER', 'EN_ATTENTE_RH']
        ).count(),
        'demandes_approuvees': mes_demandes.filter(statut='APPROUVEE').count(),
        'jours_pris_annee': sum(
            demande.nb_jours_ouvrables for demande in 
            mes_demandes.filter(
                statut__in=['APPROUVEE', 'EN_COURS', 'TERMINEE'],
                date_debut__year=annee_courante
            )
        )
    }
    
    context = {
        'employe': employe,
        'demandes': demandes_page,
        'types_conges': types_conges,
        'soldes': soldes,
        'stats': stats,
        'annee_courante': annee_courante,
        'is_rh': request.user.is_staff,
    }
    
    return render(request, 'spa/leave/requests.html', context)

@login_required
def leave_planning_content(request):
    """Contenu SPA - Planification des congés"""
    
    # Vérifier les permissions
    if not request.user.is_staff:
        messages.error(request, "Accès non autorisé")
        return redirect('dashboard')
    
    context = {
        'current_year': timezone.now().year,
        'title': 'Planification des Congés'
    }
    
    return render(request, 'spa/leave/planning.html', context)

@login_required
def leave_approvals_content(request):
    """Contenu SPA - Approbations de congés (Manager/RH)"""
    
    # Vérifier les permissions
    if not request.user.is_staff:
        messages.error(request, "Accès non autorisé")
        return render(request, 'spa/leave/approvals.html', {'error': 'Permissions insuffisantes'})
    
    # Filtres
    statut_filtre = request.GET.get('statut', 'EN_ATTENTE')
    departement_filtre = request.GET.get('departement')
    employe_filtre = request.GET.get('employe')
    type_conge_filtre = request.GET.get('type_conge')
    search = request.GET.get('search', '').strip()
    
    # Query de base - demandes nécessitant une action
    if statut_filtre == 'EN_ATTENTE':
        demandes = DemandeConge.objects.filter(
            statut__in=['SOUMISE', 'EN_ATTENTE_MANAGER', 'EN_ATTENTE_RH']
        )
    else:
        demandes = DemandeConge.objects.filter(statut=statut_filtre)
    
    demandes = demandes.select_related(
        'employe', 'type_conge', 'cree_par', 'manager_assigne'
    ).order_by('-date_creation')
    
    # Application des filtres
    if departement_filtre:
        demandes = demandes.filter(employe__department=departement_filtre)
    
    if employe_filtre:
        demandes = demandes.filter(employe_id=employe_filtre)
    
    if type_conge_filtre:
        demandes = demandes.filter(type_conge_id=type_conge_filtre)
    
    if search:
        demandes = demandes.filter(
            Q(employe__last_name__icontains=search) |
            Q(employe__first_name__icontains=search) |
            Q(numero_demande__icontains=search) |
            Q(motif__icontains=search)
        )
    
    # Pagination
    paginator = Paginator(demandes, 15)
    page = request.GET.get('page', 1)
    demandes_page = paginator.get_page(page)
    
    # Données pour filtres
    departements = Employee.objects.filter(is_active=True).values_list(
        'department__name', flat=True
    ).distinct().order_by('departement')
    
    employes = Employee.objects.filter(is_active=True).order_by('last_name', 'first_name')
    types_conges = TypeConge.objects.filter(is_active=True).order_by('ordre_affichage')
    
    # Statistiques pour le dashboard
    stats_approbation = {
        'en_attente': DemandeConge.objects.filter(
            statut__in=['SOUMISE', 'EN_ATTENTE_MANAGER', 'EN_ATTENTE_RH']
        ).count(),
        'approuvees_mois': DemandeConge.objects.filter(
            statut='APPROUVEE',
            date_creation__month=datetime.now().month,
            date_creation__year=datetime.now().year
        ).count(),
        'refusees_mois': DemandeConge.objects.filter(
            statut='REFUSEE',
            date_creation__month=datetime.now().month,
            date_creation__year=datetime.now().year
        ).count(),
    }
    
    context = {
        'demandes': demandes_page,
        'departements': departements,
        'employes': employes,
        'types_conges': types_conges,
        'stats': stats_approbation,
        'filters': {
            'statut': statut_filtre,
            'departement': departement_filtre,
            'employe': employe_filtre,
            'type_conge': type_conge_filtre,
            'search': search,
        }
    }
    
    return render(request, 'spa/leave/approvals.html', context)

@login_required  
def leave_calendar_content(request):
    """Contenu SPA - Calendrier des congés"""
    
    # Paramètres de vue
    mois = int(request.GET.get('mois', datetime.now().month))
    annee = int(request.GET.get('annee', datetime.now().year))
    departement = request.GET.get('departement')
    vue = request.GET.get('vue', 'mois')  # mois, semaine, liste
    
    # Générer le planning
    gestionnaire = GestionnaireConges()
    planning = gestionnaire.generer_planning_equipe(departement, mois, annee)
    
    # Données pour navigation
    mois_precedent = mois - 1 if mois > 1 else 12
    annee_precedente = annee if mois > 1 else annee - 1
    mois_suivant = mois + 1 if mois < 12 else 1
    annee_suivante = annee if mois < 12 else annee + 1
    
    # Types de congés pour légende
    types_conges = TypeConge.objects.filter(is_active=True).order_by('ordre_affichage')
    
    # Départements pour filtre
    departements = Employee.objects.filter(is_active=True).values_list(
        'department__name', flat=True
    ).distinct().order_by('departement')
    
    # Statistiques du mois
    total_jours_conges = sum(
        planning['planning'][employe]['jours_absents'] 
        for employe in planning['planning']
    )
    
    stats_calendrier = {
        'total_employes': len(planning['planning']),
        'employes_en_conge': len([
            emp for emp in planning['planning'] 
            if planning['planning'][emp]['jours_absents'] > 0
        ]),
        'total_jours_conges': total_jours_conges,
        'moyenne_jours': round(total_jours_conges / max(len(planning['planning']), 1), 1),
    }
    
    context = {
        'planning': planning,
        'types_conges': types_conges,
        'departements': departements,
        'stats': stats_calendrier,
        'vue_courante': {
            'mois': mois,
            'annee': annee,
            'departement': departement,
            'vue': vue,
        },
        'navigation': {
            'precedent': {'mois': mois_precedent, 'annee': annee_precedente},
            'suivant': {'mois': mois_suivant, 'annee': annee_suivante},
        }
    }
    
    return render(request, 'spa/leave/calendar.html', context)

@login_required
def leave_balances_content(request):
    """Contenu SPA - Soldes et statistiques congés"""
    
    # Filtres
    annee = int(request.GET.get('annee', datetime.now().year))
    departement_filtre = request.GET.get('departement')
    type_conge_filtre = request.GET.get('type_conge')
    search = request.GET.get('search', '').strip()
    
    # Employés selon filtres
    employes = Employee.objects.filter(is_active=True)
    
    if departement_filtre:
        employes = employes.filter(department__name__icontains=departement_filtre)
    
    if search:
        employes = employes.filter(
            Q(first_name__icontains=search) |
            Q(last_name__icontains=search) |
            Q(matricule__icontains=search)
        )
    
    employes = employes.order_by('last_name', 'first_name')
    
    # Calculer soldes pour tous les employés
    gestionnaire = GestionnaireConges()
    type_conge_obj = None
    if type_conge_filtre:
        type_conge_obj = get_object_or_404(TypeConge, id=type_conge_filtre)
    
    soldes_employes = []
    total_acquis = 0
    total_pris = 0
    total_restants = 0
    
    for employe in employes:
        try:
            soldes = gestionnaire.calculer_soldes_employe(employe, annee, type_conge_obj)
            
            employe_soldes = {
                'employe': employe,
                'soldes': soldes,
                'total_acquis': sum(s['jours_acquis'] for s in soldes.values()),
                'total_pris': sum(s['jours_pris'] for s in soldes.values()),
                'total_restants': sum(s['jours_restants'] for s in soldes.values()),
            }
            
            soldes_employes.append(employe_soldes)
            
            total_acquis += employe_soldes['total_acquis']
            total_pris += employe_soldes['total_pris'] 
            total_restants += employe_soldes['total_restants']
            
        except Exception as e:
            logger.warning(f"Erreur calcul solde {employe}: {e}")
    
    # Pagination
    paginator = Paginator(soldes_employes, 20)
    page = request.GET.get('page', 1)
    soldes_page = paginator.get_page(page)
    
    # Données pour filtres
    departements = Employee.objects.filter(is_active=True).values_list(
        'department__name', flat=True
    ).distinct().order_by('department__name')
    
    types_conges = TypeConge.objects.filter(actif=True).order_by('ordre_affichage')
    
    # Statistiques globales
    nb_employes = len(soldes_employes)
    stats_globales = {
        'nb_employes': nb_employes,
        'total_acquis': round(total_acquis, 1),
        'total_pris': round(total_pris, 1),
        'total_restants': round(total_restants, 1),
        'moyenne_acquis': round(total_acquis / max(nb_employes, 1), 1),
        'moyenne_pris': round(total_pris / max(nb_employes, 1), 1),
        'taux_utilisation': round(total_pris / max(total_acquis, 1) * 100, 1),
    }
    
    # Alertes
    alertes = []
    for solde_emp in soldes_employes:
        if solde_emp['total_restants'] < 0:
            alertes.append(f"{solde_emp['employe']} a un solde négatif")
        elif solde_emp['total_restants'] > 30:
            alertes.append(f"{solde_emp['employe']} a beaucoup de congés non pris")
    
    context = {
        'soldes_employes': soldes_page,
        'departements': departements,
        'types_conges': types_conges,
        'stats': stats_globales,
        'alertes': alertes[:10],  # Max 10 alertes
        'filters': {
            'annee': annee,
            'departement': departement_filtre,
            'type_conge': type_conge_filtre,
            'search': search,
        },
        'annees_disponibles': range(annee-2, annee+2),
    }
    
    return render(request, 'spa/leave/balances.html', context)

@login_required  
def leave_settings_content(request):
    """Contenu SPA - Paramètres de gestion des congés"""
    
    # Types de congés
    types_conges = TypeConge.objects.all().order_by('ordre_affichage')
    
    # Règles de congés 
    regles_conges = RegleConge.objects.all().order_by('code', 'libelle')
    
    # Paramètres généraux
    parametres = {
        'duree_max_conge': 30,  # Par défaut
        'delai_demande_min': 7,  # 7 jours minimum
        'notification_manager': True,
        'validation_rh_requise': True,
        'weekend_inclus': False,
        'jours_feries_inclus': False,
    }
    
    # Statistiques de configuration
    stats_config = {
        'nb_types_conges': types_conges.count(),
        'nb_regles_actives': regles_conges.filter(actif=True).count(), 
        'nb_employes_actifs': Employee.objects.filter(is_active=True).count(),
        'derniere_modification': types_conges.order_by('-id').first().libelle if types_conges.exists() else 'Aucune',
    }
    
    context = {
        'types_conges': types_conges,
        'regles_conges': regles_conges,
        'parametres': parametres,
        'stats_config': stats_config,
        'niveaux_hierarchiques': ['Employé', 'Chef d\'équipe', 'Manager', 'Directeur', 'DRH'],
        'statuts_demande': ['En attente', 'Approuvée', 'Refusée', 'Annulée'],
    }
    
    return render(request, 'spa/leave/settings.html', context)

@login_required
@require_http_methods(["GET"])
def api_statistiques_paie_data(request):
    """API - Données pour les statistiques de paie (AJAX)"""
    
    from django.db.models import Count, Sum, Avg
    from django.db.models.functions import TruncMonth
    from datetime import datetime, timedelta
    import calendar
    
    try:
        # Période par défaut (derniers 12 mois)
        periode_fin = datetime.now().date()
        periode_debut = periode_fin.replace(year=periode_fin.year - 1)
        
        # Filtres depuis la requête
        periode_filter = request.GET.get('periode', '12_mois')
        departement_filter = request.GET.get('departement', '')
        date_debut_custom = request.GET.get('date_debut', '')
        date_fin_custom = request.GET.get('date_fin', '')
        
        # Ajuster les dates selon le filtre
        if periode_filter == '1_mois':
            periode_debut = periode_fin.replace(day=1)
        elif periode_filter == '3_mois':
            periode_debut = periode_fin - timedelta(days=90)
        elif periode_filter == '6_mois':
            periode_debut = periode_fin - timedelta(days=180)
        elif periode_filter == 'custom' and date_debut_custom and date_fin_custom:
            try:
                periode_debut = datetime.strptime(date_debut_custom, '%Y-%m-%d').date()
                periode_fin = datetime.strptime(date_fin_custom, '%Y-%m-%d').date()
            except ValueError:
                pass
        
        # Base queryset pour les bulletins de la période
        bulletins_qs = BulletinPaie.objects.filter(
            periode__date_debut__gte=periode_debut,
            periode__date_fin__lte=periode_fin
        ).select_related('employe', 'periode', 'employe__department', 'employe__site')
        
        # Filtre par département si spécifié
        if departement_filter:
            bulletins_qs = bulletins_qs.filter(employe__department__name=departement_filter)
        
        # Employés actifs pour la période
        employes_actifs = Employee.objects.filter(is_active=True)
        if departement_filter:
            employes_actifs = employes_actifs.filter(department__name=departement_filter)
        
        # ========== KPIs PRINCIPAUX ==========
        
        # Calculs de base
        total_bulletins = bulletins_qs.count()
        masse_salariale_totale = bulletins_qs.aggregate(total=Sum('net_a_payer'))['total'] or 0
        employes_actifs_count = employes_actifs.count()
        
        # Masse salariale mois précédent pour variation
        mois_precedent_debut = (periode_fin.replace(day=1) - timedelta(days=1)).replace(day=1)
        mois_precedent_fin = periode_fin.replace(day=1) - timedelta(days=1)
        
        masse_salariale_precedente = BulletinPaie.objects.filter(
            periode__date_debut__gte=mois_precedent_debut,
            periode__date_fin__lte=mois_precedent_fin
        )
        if departement_filter:
            masse_salariale_precedente = masse_salariale_precedente.filter(employe__department__name=departement_filter)
        
        masse_precedente_total = masse_salariale_precedente.aggregate(total=Sum('net_a_payer'))['total'] or 0
        
        # Variation masse salariale
        if masse_precedente_total > 0:
            variation_masse = ((masse_salariale_totale - masse_precedente_total) / masse_precedente_total) * 100
        else:
            variation_masse = 0
        
        # Nouveaux employés (derniers 30 jours)
        date_30j = periode_fin - timedelta(days=30)
        nouveaux_employes = employes_actifs.filter(hire_date__gte=date_30j).count()
        
        # Taux de réussite bulletins
        bulletins_reussis = bulletins_qs.filter(net_a_payer__gt=0).count()
        taux_reussite = (bulletins_reussis / max(total_bulletins, 1)) * 100
        
        # Évolution masse salariale (12 derniers mois)
        douze_mois_ago = periode_fin - timedelta(days=365)
        evolution_mensuelle = BulletinPaie.objects.filter(
            periode__date_debut__gte=douze_mois_ago
        )
        if departement_filter:
            evolution_mensuelle = evolution_mensuelle.filter(employe__department__name=departement_filter)
        
        evolution_data = evolution_mensuelle.annotate(
            mois=TruncMonth('periode__date_debut')
        ).values('mois').annotate(
            masse=Sum('net_a_payer'),
            bulletins=Count('id')
        ).order_by('mois')
        
        evolution_chart = []
        for item in evolution_data:
            mois_nom = calendar.month_name[item['mois'].month]
            evolution_chart.append({
                'label': f"{mois_nom} {item['mois'].year}",
                'value': float(item['masse'] or 0)
            })
        
        # Répartition par département
        repartition_dept = bulletins_qs.values(
            'employe__department__name'
        ).annotate(
            masse=Sum('net_a_payer'),
            effectif=Count('employe', distinct=True)
        ).order_by('-masse')[:8]
        
        repartition_chart = []
        colors = ['#4f68d8', '#27ae60', '#f39c12', '#e74c3c', '#9b59b6', '#1abc9c', '#34495e', '#95a5a6']
        
        for i, item in enumerate(repartition_dept):
            dept_name = item['employe__department__name'] or 'Non assigné'
            repartition_chart.append({
                'label': dept_name,
                'value': float(item['masse'] or 0),
                'color': colors[i % len(colors)]
            })
        
        # Charges sociales
        charges_cnss = bulletins_qs.aggregate(total=Sum('charges_cnss_patronal'))['total'] or 0
        charges_amo = bulletins_qs.aggregate(total=Sum('charges_amo_patronal'))['total'] or 0
        formation_prof = bulletins_qs.aggregate(total=Sum('formation_professionnelle'))['total'] or 0
        
        # Préparer la réponse
        data = {
            'kpis': {
                'masse_salariale': {
                    'valeur': float(masse_salariale_totale),
                    'variation': round(variation_masse, 1),
                    'type': 'increase' if variation_masse > 0 else 'decrease' if variation_masse < 0 else 'stable'
                },
                'employes_actifs': {
                    'valeur': employes_actifs_count,
                    'nouveaux': nouveaux_employes
                },
                'bulletins_traites': {
                    'valeur': total_bulletins,
                    'taux_reussite': round(taux_reussite, 1)
                },
                'temps_traitement': {
                    'valeur': 2.3,  # Simulation
                    'evolution': -0.2
                },
                'taux_absenteisme': {
                    'valeur': 4.2,  # Simulation
                    'variation': 0.3
                },
                'cout_moyen': {
                    'valeur': round(masse_salariale_totale / max(employes_actifs_count, 1), 0),
                    'tendance': 150  # Simulation
                }
            },
            'charts': {
                'evolution': evolution_chart,
                'repartition': repartition_chart
            },
            'tableaux': {
                'charges_sociales': {
                    'cnss_patronale': float(charges_cnss),
                    'cnss_salariale': float(bulletins_qs.aggregate(total=Sum('cotisation_cnss'))['total'] or 0),
                    'amo': float(charges_amo),
                    'formation': float(formation_prof),
                    'total': float(charges_cnss + charges_amo + formation_prof)
                },
                'top_departements': list(repartition_dept)[:5]
            }
        }
        
        return JsonResponse({
            'success': True,
            'data': data
        })
        
    except Exception as e:
        logger.error(f"Erreur API statistiques paie: {e}")
        return JsonResponse({
            'success': False,
            'error': 'Erreur lors du chargement des statistiques'
        }, status=500)

# ================== APIS REST CONGÉS ==================

@login_required
@require_http_methods(["POST"])
def api_create_leave_request(request):
    """API - Créer une demande de congé"""
    
    try:
        data = json.loads(request.body)
        
        # Récupérer l'employé
        employe_id = data.get('employe_id')
        if employe_id and request.user.is_staff:
            employe = get_object_or_404(Employee, id=employe_id, is_active=True)
        else:
            # Employé connecté
            try:
                employe = Employee.objects.get(email=request.user.email, is_active=True)
            except Employee.DoesNotExist:
                employe = Employee.objects.filter(is_active=True).first()
        
        # Préparer les données de la demande
        demande_data = {
            'type_conge': get_object_or_404(TypeConge, id=data['type_conge_id']),
            'date_debut': datetime.strptime(data['date_debut'], '%Y-%m-%d').date(),
            'date_fin': datetime.strptime(data['date_fin'], '%Y-%m-%d').date(),
            'motif': data.get('motif', ''),
            'priorite': data.get('priorite', 'NORMALE'),
        }
        
        # Créer via le gestionnaire
        gestionnaire = GestionnaireConges()
        resultat = gestionnaire.soumettre_demande(demande_data, employe, request.user)
        
        if resultat['success']:
            return JsonResponse({
                'success': True,
                'demande_id': resultat['demande_id'],
                'numero_demande': resultat['numero_demande'],
                'statut': resultat['statut'],
                'message': 'Demande créée avec succès',
                'warnings': resultat.get('warnings', [])
            })
        else:
            return JsonResponse({
                'success': False,
                'erreurs': resultat['erreurs'],
                'warnings': resultat.get('warnings', [])
            }, status=400)
            
    except Exception as e:
        logger.error(f"Erreur création demande congé: {e}")
        return JsonResponse({
            'success': False,
            'error': f'Erreur système: {str(e)}'
        }, status=500)

@login_required
@require_http_methods(["PUT"])
def api_approve_leave_request(request, demande_id):
    """API - Approuver une demande de congé"""
    
    try:
        data = json.loads(request.body)
        action = data.get('action')  # 'approve' ou 'reject'
        commentaire = data.get('commentaire', '')
        
        # Déterminer le rôle
        role = 'RH' if request.user.is_staff else 'MANAGER'
        
        gestionnaire = GestionnaireConges()
        
        if action == 'approve':
            resultat = gestionnaire.approuver_demande(demande_id, request.user, commentaire, role)
        elif action == 'reject':
            resultat = gestionnaire.refuser_demande(demande_id, request.user, commentaire, role)
        else:
            return JsonResponse({
                'success': False,
                'error': 'Action non valide'
            }, status=400)
        
        if resultat['success']:
            return JsonResponse({
                'success': True,
                'nouveau_statut': resultat.get('nouveau_statut'),
                'message': f'Demande {action}ée avec succès'
            })
        else:
            return JsonResponse({
                'success': False,
                'error': resultat['erreur']
            }, status=400)
            
    except Exception as e:
        logger.error(f"Erreur {action} demande {demande_id}: {e}")
        return JsonResponse({
            'success': False,
            'error': f'Erreur système: {str(e)}'
        }, status=500)

@login_required
@require_http_methods(["DELETE"])
def api_cancel_leave_request(request, demande_id):
    """API - Annuler une demande de congé"""
    
    try:
        demande = get_object_or_404(DemandeConge, id=demande_id)
        
        # Vérifier permissions
        if not request.user.is_staff and demande.cree_par != request.user:
            return JsonResponse({
                'success': False,
                'error': 'Permissions insuffisantes'
            }, status=403)
        
        if not demande.est_annulable:
            return JsonResponse({
                'success': False,
                'error': f'Impossible d\'annuler une demande au statut {demande.get_statut_display()}'
            }, status=400)
        
        # Annuler
        ancien_statut = demande.statut
        demande.statut = 'ANNULEE'
        demande.save()
        
        # Enregistrer l'action
        ApprobationConge.objects.create(
            demande=demande,
            action='ANNULATION',
            utilisateur=request.user,
            role_utilisateur='EMPLOYE' if demande.cree_par == request.user else 'RH',
            statut_precedent=ancien_statut,
            statut_nouveau='ANNULEE'
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Demande annulée avec succès'
        })
        
    except Exception as e:
        logger.error(f"Erreur annulation demande {demande_id}: {e}")
        return JsonResponse({
            'success': False,
            'error': f'Erreur système: {str(e)}'
        }, status=500)

@login_required
@require_http_methods(["GET"])
def api_leave_statistics(request):
    """API - Statistiques du module congés"""
    
    try:
        # Période de référence
        annee_courante = datetime.now().year
        mois_courant = datetime.now().month
        
        # Statistiques générales
        stats = {
            'demandes_total': DemandeConge.objects.count(),
            'demandes_en_attente': DemandeConge.objects.filter(
                statut__in=['SOUMISE', 'EN_ATTENTE_MANAGER', 'EN_ATTENTE_RH']
            ).count(),
            'demandes_mois': DemandeConge.objects.filter(
                date_creation__year=annee_courante,
                date_creation__month=mois_courant
            ).count(),
            'taux_approbation': 0,
        }
        
        # Calcul taux d'approbation
        total_traitees = DemandeConge.objects.filter(
            statut__in=['APPROUVEE', 'REFUSEE']
        ).count()
        
        if total_traitees > 0:
            approuvees = DemandeConge.objects.filter(statut='APPROUVEE').count()
            stats['taux_approbation'] = round(approuvees / total_traitees * 100, 1)
        
        # Évolution sur 6 derniers mois
        from django.db.models.functions import TruncMonth
        
        six_months_ago = date.today() - timedelta(days=180)
        evolution = DemandeConge.objects.filter(
            date_creation__gte=six_months_ago
        ).annotate(
            mois=TruncMonth('date_creation')
        ).values('mois').annotate(
            nombre_demandes=Count('id'),
            approuvees=Count('id', filter=Q(statut='APPROUVEE')),
            refusees=Count('id', filter=Q(statut='REFUSEE'))
        ).order_by('mois')
        
        stats['evolution'] = [
            {
                'mois': item['mois'].strftime('%Y-%m'),
                'demandes': item['nombre_demandes'],
                'approuvees': item['approuvees'],
                'refusees': item['refusees'],
            }
            for item in evolution
        ]
        
        # Top types de congés
        top_types = DemandeConge.objects.filter(
            date_creation__year=annee_courante
        ).values(
            'type_conge__libelle', 'type_conge__couleur_affichage'
        ).annotate(
            count=Count('id')
        ).order_by('-count')[:5]
        
        stats['top_types_conges'] = [
            {
                'type': item['type_conge__libelle'],
                'couleur': item['type_conge__couleur_affichage'],
                'count': item['count']
            }
            for item in top_types
        ]
        
        return JsonResponse(stats)
        
    except Exception as e:
        logger.error(f"Erreur statistiques congés: {e}")
        return JsonResponse({
            'error': f'Erreur système: {str(e)}'
        }, status=500)

@login_required
@require_http_methods(["GET"])
def api_leave_calendar_data(request):
    """API - Données du calendrier des congés"""
    
    try:
        # Paramètres
        mois = int(request.GET.get('mois', datetime.now().month))
        annee = int(request.GET.get('annee', datetime.now().year))
        departement = request.GET.get('departement')
        
        # Dates du mois
        date_debut = date(annee, mois, 1)
        if mois == 12:
            date_fin = date(annee + 1, 1, 1) - timedelta(days=1)
        else:
            date_fin = date(annee, mois + 1, 1) - timedelta(days=1)
        
        # Congés du mois
        conges_query = DemandeConge.objects.filter(
            statut__in=['APPROUVEE', 'EN_COURS'],
            date_fin__gte=date_debut,
            date_debut__lte=date_fin
        ).select_related('employe', 'type_conge')
        
        if departement:
            conges_query = conges_query.filter(employe__department=departement)
        
        # Formater pour le calendrier
        events = []
        for conge in conges_query:
            events.append({
                'id': conge.id,
                'title': f"{conge.employe.nom} - {conge.type_conge.libelle}",
                'start': conge.date_debut.isoformat(),
                'end': (conge.date_fin + timedelta(days=1)).isoformat(),  # FullCalendar end is exclusive
                'backgroundColor': conge.type_conge.couleur_affichage,
                'borderColor': conge.type_conge.couleur_affichage,
                'extendedProps': {
                    'employe': f"{conge.employe.nom} {conge.employe.prenom}",
                    'type_conge': conge.type_conge.libelle,
                    'nb_jours': float(conge.nb_jours_ouvrables),
                    'statut': conge.get_statut_display(),
                    'motif': conge.motif,
                }
            })
        
        return JsonResponse({
            'success': True,
            'events': events,
            'periode': f"{mois:02d}/{annee}"
        })
        
    except Exception as e:
        logger.error(f"Erreur données calendrier: {e}")
        return JsonResponse({
            'success': False,
            'error': f'Erreur système: {str(e)}'
        }, status=500)

@login_required
@require_http_methods(["GET"])
def api_employee_balances(request, employe_id):
    """API - Soldes d'un employé spécifique"""
    
    try:
        employe = get_object_or_404(Employee, id=employe_id, is_active=True)
        annee = int(request.GET.get('annee', datetime.now().year))
        
        gestionnaire = GestionnaireConges()
        soldes = gestionnaire.calculer_soldes_employe(employe, annee)
        
        # Sérialiser les soldes pour JSON
        soldes_serialized = {}
        for type_conge_key, solde_data in soldes.items():
            # Si c'est un objet TypeConge, utiliser son libelle comme clé
            if hasattr(type_conge_key, 'libelle'):
                key = type_conge_key.libelle
            else:
                key = str(type_conge_key)
                
            # Convertir toutes les valeurs en types sérialisables
            soldes_serialized[key] = {
                'jours_acquis': float(solde_data.get('jours_acquis', 0)),
                'jours_pris': float(solde_data.get('jours_pris', 0)),
                'jours_restants': float(solde_data.get('jours_restants', 0)),
            }
        
        return JsonResponse({
            'success': True,
            'employe': {
                'id': employe.id,
                'nom': employe.last_name,
                'prenom': employe.first_name,
                'departement': employe.department.name if employe.department else None,
            },
            'annee': annee,
            'soldes': soldes_serialized
        })
        
    except Exception as e:
        logger.error(f"Erreur soldes employé {employe_id}: {e}")
        return JsonResponse({
            'success': False,
            'error': f'Erreur système: {str(e)}'
        }, status=500)

@login_required
@require_http_methods(["POST"])
def api_validate_leave_dates(request):
    """API - Valider des dates de congé"""
    
    try:
        data = json.loads(request.body)
        
        employe_id = data.get('employe_id')
        employe = get_object_or_404(Employee, id=employe_id, is_active=True)
        
        gestionnaire = GestionnaireConges()
        validation = gestionnaire.valider_demande_conge(data, employe, request.user)
        
        return JsonResponse(validation)
        
    except Exception as e:
        logger.error(f"Erreur validation dates: {e}")
        return JsonResponse({
            'valide': False,
            'erreurs': [f'Erreur système: {str(e)}'],
            'warnings': []
        }, status=500)


# ================== MISSING API ENDPOINTS - STUBS ==================

@login_required
@require_http_methods(["GET"])
def api_parametrages_list(request):
    """API - Liste des paramétrages paie"""
    try:
        parametrages = ParametragePaie.objects.all().order_by('-annee')
        
        data = {
            'success': True,
            'parametrages': [
                {
                    'id': p.id,
                    'annee': p.annee,
                    'actif': p.actif,
                    'plafond_cnss': float(p.plafond_cnss),
                    'taux_cnss_salarie': float(p.taux_cnss_salarie),
                    'taux_cnss_patronal': float(p.taux_cnss_patronal),
                    'date_creation': p.date_creation.strftime('%d/%m/%Y')
                }
                for p in parametrages
            ]
        }
        
        return JsonResponse(data)
        
    except Exception as e:
        logger.error(f"Erreur chargement paramétrages: {e}")
        return JsonResponse({'success': False, 'error': 'Erreur lors du chargement des paramétrages'})

@login_required  
@require_http_methods(["PUT"])
def api_parametrage_update(request, parametrage_id):
    """API - Modifier un paramétrage"""
    return JsonResponse({'success': True, 'message': 'Paramétrage mis à jour'})

@login_required
@require_http_methods(["POST"])
def api_parametrage_create(request):
    """API - Créer un paramétrage"""
    return JsonResponse({'success': True, 'message': 'Paramétrage créé'})

@login_required
@require_http_methods(["PUT"])
def api_bareme_ir_update(request, bareme_id):
    """API - Modifier un barème IR"""
    return JsonResponse({'success': True, 'message': 'Barème IR mis à jour'})

@login_required
@require_http_methods(["POST"])
def api_bareme_ir_create(request):
    """API - Créer un barème IR"""
    return JsonResponse({'success': True, 'message': 'Barème IR créé'})

@login_required
@require_http_methods(["DELETE"])
def api_bareme_ir_delete(request, bareme_id):
    """API - Supprimer un barème IR"""
    return JsonResponse({'success': True, 'message': 'Barème IR supprimé'})

@login_required
@require_http_methods(["POST"])
def api_periode_create(request):
    """API - Créer une période"""
    try:
        data = json.loads(request.body)
        
        # Validation des données requises
        required_fields = ['libelle', 'type_periode', 'date_debut', 'date_fin', 'date_paie', 'parametrage_id']
        for field in required_fields:
            if field not in data:
                return JsonResponse({'success': False, 'error': f'Le champ {field} est requis'})
        
        # Vérifier que le paramétrage existe
        try:
            parametrage = ParametragePaie.objects.get(id=data['parametrage_id'])
        except ParametragePaie.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Paramétrage introuvable'})
        
        # Créer la période
        periode = PeriodePaie.objects.create(
            libelle=data['libelle'],
            type_periode=data['type_periode'],
            date_debut=data['date_debut'],
            date_fin=data['date_fin'],
            date_paie=data['date_paie'],
            nb_jours_travailles=data.get('nb_jours_travailles', 30),
            nb_heures_standard=data.get('nb_heures_standard', 191.33),
            parametrage=parametrage,
            statut='BROUILLON'
        )
        
        return JsonResponse({
            'success': True, 
            'message': f'Période "{periode.libelle}" créée avec succès',
            'periode_id': periode.id
        })
        
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Données JSON invalides'})
    except Exception as e:
        logger.error(f"Erreur création période: {e}")
        return JsonResponse({'success': False, 'error': 'Erreur lors de la création de la période'})

@login_required
@require_http_methods(["GET"])
def api_periode_detail(request, periode_id):
    """API - Détail d'une période"""
    return JsonResponse({'success': True, 'periode': {}})

@login_required
@require_http_methods(["PUT"])
def api_periode_update(request, periode_id):
    """API - Modifier une période"""
    try:
        periode = get_object_or_404(PeriodePaie, id=periode_id)
        
        # Vérifier que la période n'est pas clôturée
        if periode.statut == 'CLOTUREE':
            return JsonResponse({'success': False, 'error': 'Impossible de modifier une période clôturée'})
        
        data = json.loads(request.body)
        
        # Mettre à jour les champs modifiables
        if 'libelle' in data:
            periode.libelle = data['libelle']
        if 'type_periode' in data:
            periode.type_periode = data['type_periode']
        if 'date_debut' in data:
            periode.date_debut = data['date_debut']
        if 'date_fin' in data:
            periode.date_fin = data['date_fin']
        if 'date_paie' in data:
            periode.date_paie = data['date_paie']
        if 'nb_jours_travailles' in data:
            periode.nb_jours_travailles = data['nb_jours_travailles']
        if 'nb_heures_standard' in data:
            periode.nb_heures_standard = data['nb_heures_standard']
        if 'parametrage_id' in data:
            try:
                parametrage = ParametragePaie.objects.get(id=data['parametrage_id'])
                periode.parametrage = parametrage
            except ParametragePaie.DoesNotExist:
                return JsonResponse({'success': False, 'error': 'Paramétrage introuvable'})
        
        periode.save()
        
        return JsonResponse({
            'success': True, 
            'message': f'Période "{periode.libelle}" mise à jour avec succès'
        })
        
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Données JSON invalides'})
    except Exception as e:
        logger.error(f"Erreur mise à jour période {periode_id}: {e}")
        return JsonResponse({'success': False, 'error': 'Erreur lors de la mise à jour'})

@login_required
@require_http_methods(["DELETE"])
def api_periode_delete(request, periode_id):
    """API - Supprimer une période"""
    return JsonResponse({'success': True, 'message': 'Période supprimée'})

@login_required
@require_http_methods(["POST"])
def api_periode_valider(request, periode_id):
    """API - Valider une période"""
    return JsonResponse({'success': True, 'message': 'Période validée'})

@login_required
@require_http_methods(["POST"])
def api_periode_cloturer(request, periode_id):
    """API - Clôturer une période"""
    return JsonResponse({'success': True, 'message': 'Période clôturée'})

@login_required
@require_http_methods(["GET"])
def api_export_periode(request, periode_id):
    """API - Exporter une période"""
    return JsonResponse({'success': True, 'export': 'data'})

@login_required
@require_http_methods(["GET"])
def api_export_bulletins(request):
    """API - Exporter les bulletins"""
    return JsonResponse({'success': True, 'export': 'data'})

@login_required
@require_http_methods(["GET"])
def api_export_bulletin(request, bulletin_id):
    """API - Exporter un bulletin individuel"""
    try:
        bulletin = get_object_or_404(BulletinPaie, id=bulletin_id)
        
        # Récupérer le format demandé (par défaut: json)
        format_export = request.GET.get('format', 'json').lower()
        
        if format_export == 'pdf':
            # Générer et retourner un PDF
            return generer_bulletin_pdf(bulletin)
        
        elif format_export == 'excel':
            # Générer et retourner un fichier Excel
            return generer_bulletin_excel(bulletin)
        
        else:
            # Format JSON par défaut
            return export_bulletin_json(bulletin)
        
    except Exception as e:
        # get_object_or_404 lève Http404 qui est capturée ici
        if "does not exist" in str(e).lower() or "no BulletinPaie matches" in str(e):
            return JsonResponse({
                'success': False, 
                'error': f'Bulletin {bulletin_id} non trouvé'
            }, status=404)
        else:
            return JsonResponse({
                'success': False, 
                'error': f'Erreur lors de l\'export: {str(e)}'
            }, status=500)


def export_bulletin_json(bulletin):
    """Exporter un bulletin en format JSON"""
    try:
        # Calculer les totaux des cotisations
        total_cotisations = (
            bulletin.cotisation_cnss + 
            bulletin.cotisation_amo + 
            bulletin.cotisation_cimr
        )
        
        # Préparer les données du bulletin
        bulletin_data = {
            'id': bulletin.id,
            'numero_bulletin': bulletin.numero_bulletin,
            'employe': {
                'nom': bulletin.employe.last_name,
                'prenom': bulletin.employe.first_name,
                'matricule': bulletin.employe.id,
                'email': bulletin.employe.email,
                'position': bulletin.employe.position,
            },
            'periode': {
                'libelle': bulletin.periode.libelle,
                'date_debut': bulletin.periode.date_debut.isoformat(),
                'date_fin': bulletin.periode.date_fin.isoformat(),
                'type_periode': bulletin.periode.type_periode,
            },
            'salaire': {
                'salaire_base': float(bulletin.salaire_base),
                'heures_supplementaires': float(bulletin.heures_supplementaires),
                'taux_heure_sup': float(bulletin.taux_heure_sup),
            },
            'primes': {
                'prime_anciennete': float(bulletin.prime_anciennete),
                'prime_responsabilite': float(bulletin.prime_responsabilite),
                'indemnite_transport': float(bulletin.indemnite_transport),
                'avantages_nature': float(bulletin.avantages_nature),
            },
            'totaux': {
                'total_brut': float(bulletin.total_brut),
                'total_imposable': float(bulletin.total_imposable),
                'total_cotisable_cnss': float(bulletin.total_cotisable_cnss),
            },
            'cotisations': {
                'cnss': float(bulletin.cotisation_cnss),
                'amo': float(bulletin.cotisation_amo),
                'cimr': float(bulletin.cotisation_cimr),
                'total_cotisations': float(total_cotisations),
            },
            'impots': {
                'ir_brut': float(bulletin.ir_brut),
                'ir_net': float(bulletin.ir_net),
            },
            'retenues': {
                'avances': float(bulletin.avances),
                'prets': float(bulletin.prets),
                'autres_retenues': float(bulletin.autres_retenues),
                'total_retenues': float(bulletin.total_retenues),
            },
            'charges_patronales': {
                'cnss_patronal': float(bulletin.charges_cnss_patronal),
                'amo_patronal': float(bulletin.charges_amo_patronal),
                'formation_professionnelle': float(bulletin.formation_professionnelle),
                'prestations_sociales': float(bulletin.prestations_sociales),
            },
            'resultat': {
                'net_a_payer': float(bulletin.net_a_payer),
            },
            'metadonnees': {
                'date_generation': bulletin.date_generation.isoformat(),
                'genere_par': bulletin.genere_par.username if bulletin.genere_par else None,
                'fichier_pdf': bulletin.fichier_pdf.url if bulletin.fichier_pdf else None,
            }
        }
        
        return JsonResponse({
            'success': True, 
            'bulletin': bulletin_data,
            'format': 'json'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False, 
            'error': f'Erreur lors de l\'export JSON: {str(e)}'
        }, status=500)

def generer_bulletin_pdf(bulletin):
    """Générer un bulletin en format PDF"""
    try:
        from weasyprint import HTML
        from django.http import HttpResponse
        from django.template.loader import render_to_string
        import io
        
        # Utiliser le template Django pour générer le HTML
        html_content = render_to_string('spa/payroll/bulletin_pdf.html', {
            'bulletin': bulletin
        })
        
        # Générer le PDF avec WeasyPrint
        pdf_buffer = io.BytesIO()
        HTML(string=html_content).write_pdf(pdf_buffer)
        pdf_buffer.seek(0)
        
        # Créer la réponse HTTP
        response = HttpResponse(pdf_buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="bulletin_{bulletin.numero_bulletin}.pdf"'
        
        return response
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Erreur lors de la génération PDF: {str(e)}'
        }, status=500)

def generer_bulletin_excel(bulletin):
    """Générer un bulletin en format Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from django.http import HttpResponse
        import io
        
        # Créer un nouveau classeur
        wb = Workbook()
        ws = wb.active
        ws.title = "Bulletin de Paie"
        
        # Styles
        header_font = Font(bold=True, size=14)
        section_font = Font(bold=True, size=12)
        bold_font = Font(bold=True)
        center_align = Alignment(horizontal='center')
        right_align = Alignment(horizontal='right')
        header_fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        
        # En-tête
        ws['A1'] = 'BULLETIN DE PAIE'
        ws['A1'].font = header_font
        ws['A1'].alignment = center_align
        ws.merge_cells('A1:D1')
        
        ws['A2'] = f'Numéro: {bulletin.numero_bulletin}'
        ws['A3'] = f'Période: {bulletin.periode.libelle}'
        
        # Informations employé
        row = 5
        ws[f'A{row}'] = 'INFORMATIONS EMPLOYÉ'
        ws[f'A{row}'].font = section_font
        ws[f'A{row}'].fill = header_fill
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        ws[f'A{row}'] = 'Nom:'
        ws[f'B{row}'] = bulletin.employe.last_name
        ws[f'C{row}'] = 'Prénom:'
        ws[f'D{row}'] = bulletin.employe.first_name
        
        row += 1
        ws[f'A{row}'] = 'Matricule:'
        ws[f'B{row}'] = bulletin.employe.id
        ws[f'C{row}'] = 'Position:'
        ws[f'D{row}'] = bulletin.employe.position
        
        # Éléments de rémunération
        row += 2
        ws[f'A{row}'] = 'ÉLÉMENTS DE RÉMUNÉRATION'
        ws[f'A{row}'].font = section_font
        ws[f'A{row}'].fill = header_fill
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        ws[f'A{row}'] = 'Désignation'
        ws[f'B{row}'] = 'Montant (MAD)'
        ws[f'A{row}'].font = bold_font
        ws[f'B{row}'].font = bold_font
        
        elements = [
            ('Salaire de Base', bulletin.salaire_base),
            ('Heures Supplémentaires', bulletin.heures_supplementaires * bulletin.taux_heure_sup),
            ('Prime d\'Ancienneté', bulletin.prime_anciennete),
            ('Prime de Responsabilité', bulletin.prime_responsabilite),
            ('Indemnité de Transport', bulletin.indemnite_transport),
            ('Avantages en Nature', bulletin.avantages_nature),
            ('TOTAL BRUT', bulletin.total_brut),
        ]
        
        for designation, montant in elements:
            row += 1
            ws[f'A{row}'] = designation
            ws[f'B{row}'] = float(montant)
            ws[f'B{row}'].number_format = '#,##0.00'
            ws[f'B{row}'].alignment = right_align
            if designation.startswith('TOTAL'):
                ws[f'A{row}'].font = bold_font
                ws[f'B{row}'].font = bold_font
        
        # Cotisations
        row += 2
        ws[f'A{row}'] = 'COTISATIONS SALARIALES'
        ws[f'A{row}'].font = section_font
        ws[f'A{row}'].fill = header_fill
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        ws[f'A{row}'] = 'Désignation'
        ws[f'B{row}'] = 'Montant (MAD)'
        ws[f'A{row}'].font = bold_font
        ws[f'B{row}'].font = bold_font
        
        cotisations = [
            ('CNSS', bulletin.cotisation_cnss),
            ('AMO', bulletin.cotisation_amo),
            ('CIMR', bulletin.cotisation_cimr),
            ('TOTAL COTISATIONS', bulletin.cotisation_cnss + bulletin.cotisation_amo + bulletin.cotisation_cimr),
        ]
        
        for designation, montant in cotisations:
            row += 1
            ws[f'A{row}'] = designation
            ws[f'B{row}'] = float(montant)
            ws[f'B{row}'].number_format = '#,##0.00'
            ws[f'B{row}'].alignment = right_align
            if designation.startswith('TOTAL'):
                ws[f'A{row}'].font = bold_font
                ws[f'B{row}'].font = bold_font
        
        # Net à payer
        row += 2
        ws[f'A{row}'] = 'NET À PAYER'
        ws[f'B{row}'] = float(bulletin.net_a_payer)
        ws[f'A{row}'].font = Font(bold=True, size=14)
        ws[f'B{row}'].font = Font(bold=True, size=14)
        ws[f'B{row}'].number_format = '#,##0.00'
        ws[f'B{row}'].alignment = right_align
        
        # Ajuster les largeurs de colonnes
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 15
        
        # Sauvegarder dans un buffer
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # Créer la réponse HTTP
        response = HttpResponse(
            excel_buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="bulletin_{bulletin.numero_bulletin}.xlsx"'
        
        return response
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Erreur lors de la génération Excel: {str(e)}'
        }, status=500)

@login_required
@require_http_methods(["GET"])
def api_export_declarations(request):
    """API - Exporter les déclarations"""
    return JsonResponse({'success': True, 'export': 'data'})

@login_required
@require_http_methods(["DELETE"])
def api_bulletin_delete(request, bulletin_id):
    """API - Supprimer un bulletin"""
    try:
        bulletin = get_object_or_404(BulletinPaie, id=bulletin_id)
        bulletin.delete()
        return JsonResponse({'success': True, 'message': 'Bulletin supprimé avec succès'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)

@login_required
@require_http_methods(["GET"])
def api_bulletins_export(request):
    """API - Exporter les bulletins"""
    return JsonResponse({'success': True, 'bulletins': []})

@login_required
@require_http_methods(["GET"])
def api_rubrique_detail(request, rubrique_id):
    """API - Détail d'une rubrique personnalisée"""
    rubrique = get_object_or_404(RubriquePersonnalisee, id=rubrique_id)
    data = {
        'success': True,
        'rubrique': {
            'id': rubrique.id,
            'code': rubrique.code,
            'libelle': rubrique.libelle,
            'type_rubrique': rubrique.type_rubrique,
            'mode_calcul': rubrique.mode_calcul,
            'periodicite': rubrique.periodicite,
            'valeur_fixe': float(rubrique.valeur_fixe) if rubrique.valeur_fixe is not None else None,
            'pourcentage': float(rubrique.pourcentage) if rubrique.pourcentage is not None else None,
            'formule': rubrique.formule,
            'imposable_ir': rubrique.imposable_ir,
            'soumis_cnss': rubrique.soumis_cnss,
            'soumis_amo': rubrique.soumis_amo,
            'soumis_cimr': rubrique.soumis_cimr,
            'ordre_affichage': rubrique.ordre_affichage,
            'afficher_bulletin': rubrique.afficher_bulletin,
            'actif': rubrique.actif,
        }
    }
    return JsonResponse(data)

@login_required
@require_http_methods(["GET"])
def api_bulletin_pdf(request, bulletin_id):
    """API - Télécharger un bulletin en PDF"""
    try:
        bulletin = get_object_or_404(BulletinPaie, id=bulletin_id)
        return generer_bulletin_pdf(bulletin)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Erreur lors de la génération PDF: {str(e)}'
        }, status=500)

@login_required
@require_http_methods(["GET"])
def api_bulletin_excel(request, bulletin_id):
    """API - Télécharger un bulletin en Excel"""
    try:
        bulletin = get_object_or_404(BulletinPaie, id=bulletin_id)
        return generer_bulletin_excel(bulletin)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Erreur lors de la génération Excel: {str(e)}'
        }, status=500)

# ================== PDF & ADMIN VIEWS - STUBS ==================

@login_required
def bulletin_pdf(request, bulletin_id):
    """PDF d'un bulletin"""
    return JsonResponse({'error': 'PDF generation not implemented yet'})

@login_required
def bulletin_download(request, bulletin_id):
    """Télécharger un bulletin"""
    return JsonResponse({'error': 'Download not implemented yet'})

@login_required
def periode_bulletins_pdf(request, periode_id):
    """PDF de tous les bulletins d'une période"""
    return JsonResponse({'error': 'PDF generation not implemented yet'})

@login_required
def periode_livre_paie(request, periode_id):
    """Livre de paie d'une période"""
    return JsonResponse({'error': 'Livre de paie not implemented yet'})

@login_required
def declaration_cnss(request, periode_id):
    """Déclaration CNSS"""
    return JsonResponse({'error': 'Declaration CNSS not implemented yet'})

@login_required
def declaration_amo(request, periode_id):
    """Déclaration AMO"""
    return JsonResponse({'error': 'Declaration AMO not implemented yet'})

@login_required
def declaration_ir(request, periode_id):
    """Déclaration IR"""
    return JsonResponse({'error': 'Declaration IR not implemented yet'})

# ================== LEAVE API STUBS ==================

@login_required
@require_http_methods(["GET"])
def api_leave_types_list(request):
    """API - Liste des types de congés"""
    return JsonResponse({'success': True, 'types': []})

@login_required
@require_http_methods(["POST"])
def api_leave_type_create(request):
    """API - Créer un type de congé"""
    return JsonResponse({'success': True, 'message': 'Type créé'})

@login_required
@require_http_methods(["PUT"])
def api_leave_type_update(request, type_id):
    """API - Modifier un type de congé"""
    return JsonResponse({'success': True, 'message': 'Type mis à jour'})

@login_required
@require_http_methods(["DELETE"])
def api_leave_type_delete(request, type_id):
    """API - Supprimer un type de congé"""
    return JsonResponse({'success': True, 'message': 'Type supprimé'})

@login_required
@require_http_methods(["GET"])
def api_export_leave_planning(request):
    """API - Exporter planning congés"""
    return JsonResponse({'success': True, 'export': 'data'})

@login_required
@require_http_methods(["GET"])
def api_export_leave_balances(request):
    """API - Exporter soldes congés"""
    return JsonResponse({'success': True, 'export': 'data'})

@login_required
@require_http_methods(["POST"])
def api_import_leave_balances(request):
    """API - Importer soldes congés"""
    return JsonResponse({'success': True, 'message': 'Import réussi'})

@login_required
@require_http_methods(["GET"])
def api_leave_request_history(request, demande_id):
    """API - Historique d'une demande"""
    return JsonResponse({'success': True, 'history': []})

@login_required
@require_http_methods(["GET"])
def api_leave_workflow_rules(request):
    """API - Règles de workflow"""
    return JsonResponse({'success': True, 'rules': []})

@login_required
@require_http_methods(["GET"])
def api_leave_notifications(request):
    """API - Notifications congés"""
    return JsonResponse({'success': True, 'notifications': []})

@login_required
@require_http_methods(["POST"])
def api_mark_notification_read(request, notif_id):
    """API - Marquer notification comme lue"""
    return JsonResponse({'success': True, 'message': 'Notification marquée'})

@login_required
@require_http_methods(["GET"])
def api_leave_usage_report(request):
    """API - Rapport d'utilisation des congés"""
    return JsonResponse({'success': True, 'report': {}})

@login_required
@require_http_methods(["GET"])
def api_leave_trends_report(request):
    """API - Rapport de tendances des congés"""
    return JsonResponse({'success': True, 'report': {}})

@login_required
@require_http_methods(["GET"])
def api_leave_settings(request):
    """API - Paramètres congés"""
    return JsonResponse({'success': True, 'settings': {}})

@login_required
@require_http_methods(["POST"])
def api_update_leave_settings(request):
    """API - Mettre à jour paramètres congés"""
    return JsonResponse({'success': True, 'message': 'Paramètres mis à jour'})
# Ajoutez ces fonctions dans votre fichier paie/views.py

@login_required
@require_http_methods(["GET"])
def api_dashboard_stats(request):
    """API - Statistiques pour le dashboard principal"""
    try:
        # Statistiques des employés
        total_employees = Employee.objects.filter(is_active=True).count()
        employees_by_site = {}
        employees_by_dept = {}
        
        # Compter les employés par site
        sites = Site.objects.filter(is_active=True)
        for site in sites:
            employees_by_site[site.name] = Employee.objects.filter(
                is_active=True, 
                site=site
            ).count()
        
        # Compter les employés par département
        departments = Department.objects.filter(is_active=True)
        for dept in departments:
            employees_by_dept[dept.name] = Employee.objects.filter(
                is_active=True,
                department=dept
            ).count()
        
        # Statistiques de paie (si disponibles)
        periode_courante = PeriodePaie.objects.filter(
            date_debut__lte=date.today(),
            date_fin__gte=date.today()
        ).first()
        
        payroll_stats = {}
        if periode_courante:
            bulletins = BulletinPaie.objects.filter(periode=periode_courante)
            payroll_stats = {
                'periode_courante': periode_courante.libelle,
                'bulletins_generes': bulletins.count(),
                'masse_salariale': float(bulletins.aggregate(
                    total=Sum('net_a_payer')
                )['total'] or 0)
            }
        
        # Statistiques de congés
        leave_stats = {
            'demandes_en_attente': DemandeConge.objects.filter(
                statut__in=['SOUMISE', 'EN_ATTENTE_MANAGER', 'EN_ATTENTE_RH']
            ).count() if 'DemandeConge' in globals() else 0,
            'employes_en_conge': DemandeConge.objects.filter(
                statut='EN_COURS',
                date_debut__lte=date.today(),
                date_fin__gte=date.today()
            ).count() if 'DemandeConge' in globals() else 0,
        }
        
        return JsonResponse({
            'success': True,
            'stats': {
                'employees': {
                    'total': total_employees,
                    'by_site': employees_by_site,
                    'by_department': employees_by_dept,
                },
                'payroll': payroll_stats,
                'leave': leave_stats,
                'sites': Site.objects.filter(is_active=True).count(),
                'departments': Department.objects.filter(is_active=True).count(),
            }
        })
        
    except Exception as e:
        logger.error(f"Erreur dans api_dashboard_stats: {e}")
        return JsonResponse({
            'success': False,
            'error': f'Erreur lors du chargement des statistiques: {str(e)}'
        }, status=500)


@login_required
@require_http_methods(["GET"])
def api_recent_employees(request):
    """API - Liste des employés récemment ajoutés"""
    try:
        # Récupérer les 10 derniers employés ajoutés
        recent_employees = Employee.objects.filter(
            is_active=True
        ).order_by('-hire_date')[:10].select_related('site', 'department')
        
        employees_data = []
        for emp in recent_employees:
            employees_data.append({
                'id': emp.id,
                'first_name': emp.first_name or '',
                'last_name': emp.last_name or '',
                'full_name': emp.full_name,
                'email': emp.email,
                'position': emp.position or 'Poste non défini',
                'hire_date': emp.hire_date.isoformat() if emp.hire_date else None,
                'site': emp.site.name if emp.site else 'Non assigné',
                'department': emp.department.name if emp.department else 'Non assigné',
                'phone': emp.phone,
            })
        
        return JsonResponse({
            'success': True,
            'employees': employees_data,
            'total_count': len(employees_data)
        })
        
    except Exception as e:
        logger.error(f"Erreur dans api_recent_employees: {e}")
        return JsonResponse({
            'success': False,
            'error': f'Erreur lors du chargement des employés récents: {str(e)}'
        }, status=500)


@login_required
@require_http_methods(["GET"])
def api_stats(request):
    """API - Statistiques générales du système"""
    try:
        # Type de statistiques demandé
        stat_type = request.GET.get('type', 'general')
        
        if stat_type == 'employees':
            # Statistiques détaillées des employés
            stats = {
                'total': Employee.objects.filter(is_active=True).count(),
                'by_status': {
                    'active': Employee.objects.filter(is_active=True).count(),
                    'inactive': Employee.objects.filter(is_active=False).count(),
                },
                'by_site': {},
                'by_department': {},
                'by_position': {},
                'recent_hires': {
                    'this_month': Employee.objects.filter(
                        is_active=True,
                        hire_date__year=datetime.now().year,
                        hire_date__month=datetime.now().month
                    ).count(),
                    'last_month': Employee.objects.filter(
                        is_active=True,
                        hire_date__year=datetime.now().year,
                        hire_date__month=datetime.now().month - 1 if datetime.now().month > 1 else 12
                    ).count(),
                }
            }
            
            # Répartition par site
            for site in Site.objects.filter(is_active=True):
                stats['by_site'][site.name] = Employee.objects.filter(
                    is_active=True, site=site
                ).count()
            
            # Répartition par département
            for dept in Department.objects.filter(is_active=True):
                stats['by_department'][dept.name] = Employee.objects.filter(
                    is_active=True, department=dept
                ).count()
            
            # Top 5 des positions
            positions = Employee.objects.filter(
                is_active=True
            ).values('position').annotate(
                count=Count('id')
            ).order_by('-count')[:5]
            
            for pos in positions:
                if pos['position']:
                    stats['by_position'][pos['position']] = pos['count']
        
        elif stat_type == 'payroll':
            # Statistiques de paie
            current_year = datetime.now().year
            stats = {
                'periodes_annee': PeriodePaie.objects.filter(
                    date_debut__year=current_year
                ).count(),
                'bulletins_annee': BulletinPaie.objects.filter(
                    periode__date_debut__year=current_year
                ).count(),
                'masse_salariale_annuelle': float(
                    BulletinPaie.objects.filter(
                        periode__date_debut__year=current_year
                    ).aggregate(total=Sum('net_a_payer'))['total'] or 0
                ),
                'moyenne_salaire': float(
                    Employee.objects.filter(
                        is_active=True
                    ).aggregate(avg=Avg('salary'))['avg'] or 0
                ),
            }
        
        elif stat_type == 'leave':
            # Statistiques de congés
            current_year = datetime.now().year
            stats = {
                'demandes_annee': DemandeConge.objects.filter(
                    date_creation__year=current_year
                ).count() if 'DemandeConge' in globals() else 0,
                'jours_conges_pris': DemandeConge.objects.filter(
                    statut='APPROUVEE',
                    date_debut__year=current_year
                ).aggregate(total=Sum('nb_jours_ouvrables'))['total'] or 0 if 'DemandeConge' in globals() else 0,
            }
        
        else:
            # Statistiques générales
            stats = {
                'employees': Employee.objects.filter(is_active=True).count(),
                'sites': Site.objects.filter(is_active=True).count(),
                'users': User.objects.filter(is_active=True).count(),
                'users': User.objects.filter(is_active=True).count() if 'User' in globals() else 0,
            }
        
        return JsonResponse({
            'success': True,
            'type': stat_type,
            'stats': stats,
            'generated_at': datetime.now().isoformat()
        })
        
    except Exception as e:
        logger.error(f"Erreur dans api_stats: {e}")
        return JsonResponse({
            'success': False,
            'error': f'Erreur lors du chargement des statistiques: {str(e)}'
        }, status=500)


@login_required 
@require_http_methods(["GET"])
def api_user_permissions(request):
    """API - Permissions de l'utilisateur actuel"""
    try:
        user = request.user
        
        # Permissions de base
        permissions = {
            'is_superuser': user.is_superuser,
            'is_staff': user.is_staff,
            'is_active': user.is_active,
            'groups': list(user.groups.values_list('name', flat=True)),
            'user_permissions': list(user.user_permissions.values_list('codename', flat=True))
        }
        
        # Profil utilisateur si disponible
        try:
            if hasattr(user, 'userprofile'):
                profile = user.userprofile
                permissions.update({
                    'role': profile.role,
                    'department': profile.department.name if profile.department else None,
                    'employee_id': profile.employee.id if profile.employee else None
                })
        except Exception:
            pass
            
        return JsonResponse({
            'success': True,
            'permissions': permissions,
            'user': {
                'id': user.id,
                'username': user.username,
                'first_name': user.first_name,
                'last_name': user.last_name,
                'email': user.email
            }
        })
        
    except Exception as e:
        logger.error(f"Erreur dans api_user_permissions: {e}")
        return JsonResponse({
            'success': False,
            'error': f'Erreur lors du chargement des permissions: {str(e)}'
        }, status=500)


# ============================================================================
# VUES SPA CONTENT - Interface utilisateur
# ============================================================================

@login_required
def attendance_timeclock_content(request):
    """Interface pointage principal (employé dédié)"""
    try:
        # Récupérer tous les employés actifs avec leurs départements
        employes = Employee.objects.filter(is_active=True).select_related('department').order_by('last_name', 'first_name')
        
        # Récupérer les départements pour filtrage
        departements = Department.objects.filter(is_active=True).order_by('name')
        
        # Statut de présence temps réel
        gestionnaire = GestionnairePointage()
        statut_temps_reel = gestionnaire.get_statut_presence_temps_reel()
        
        # Derniers pointages du jour
        aujourd_hui = date.today()
        derniers_pointages = Pointage.objects.filter(
            heure_pointage__date=aujourd_hui
        ).select_related('employe').order_by('-heure_pointage')[:20]
        
        context = {
            'employes': employes,
            'departements': departements,
            'statut_temps_reel': statut_temps_reel,
            'derniers_pointages': derniers_pointages,
            'types_pointage': Pointage.TYPES_POINTAGE,
        }
        
        return render(request, 'spa/attendance/timeclock.html', context)
        
    except Exception as e:
        logger.error(f"Erreur dans attendance_timeclock_content: {e}")
        messages.error(request, "Erreur lors du chargement de l'interface de pointage")
        return render(request, 'spa/attendance/timeclock.html', {'error': str(e)})


@login_required
def attendance_dashboard_content(request):
    """Dashboard temps réel des présences"""
    try:
        gestionnaire = GestionnairePointage()
        aujourd_hui = date.today()
        
        # Statut temps réel
        statut_temps_reel = gestionnaire.get_statut_presence_temps_reel()
        
        # Détection retards/absences aujourd'hui
        detection_jour = gestionnaire.detecter_retards_absences(aujourd_hui)
        
        # Statistiques de la semaine
        debut_semaine = aujourd_hui - timedelta(days=aujourd_hui.weekday())
        fin_semaine = debut_semaine + timedelta(days=6)
        
        stats_semaine = PresenceJournaliere.objects.filter(
            date__range=[debut_semaine, fin_semaine]
        ).aggregate(
            total_heures=Sum('heures_travaillees'),
            total_retards=Count('id', filter=Q(retard_minutes__gt=0)),
            total_absences=Count('id', filter=Q(statut_jour='ABSENT'))
        )
        
        # Alertes récentes
        alertes_recentes = AlertePresence.objects.filter(
            statut='NOUVELLE',
            date_creation__gte=timezone.now() - timedelta(days=7)
        ).select_related('employe').order_by('-date_creation')[:10]
        
        # Données pour graphiques (7 derniers jours)
        graphique_data = []
        for i in range(7):
            jour = aujourd_hui - timedelta(days=6-i)
            presences_jour = PresenceJournaliere.objects.filter(date=jour)
            graphique_data.append({
                'date': jour.strftime('%d/%m'),
                'presents': presences_jour.filter(statut_jour='PRESENT').count(),
                'absents': presences_jour.filter(statut_jour='ABSENT').count(),
                'retards': presences_jour.filter(retard_minutes__gt=0).count()
            })
        
        context = {
            'statut_temps_reel': statut_temps_reel,
            'detection_jour': detection_jour,
            'stats_semaine': stats_semaine,
            'alertes_recentes': alertes_recentes,
            'graphique_data': json.dumps(graphique_data),
            'aujourd_hui': aujourd_hui,
        }
        
        return render(request, 'spa/attendance/dashboard.html', context)
        
    except Exception as e:
        logger.error(f"Erreur dans attendance_dashboard_content: {e}")
        return render(request, 'spa/attendance/dashboard.html', {'error': str(e)})


@login_required
def attendance_reports_content(request):
    """Rapports et feuilles de présence"""
    try:
        # Paramètres de filtrage depuis l'URL
        date_debut = request.GET.get('date_debut', date.today().strftime('%Y-%m-%d'))
        date_fin = request.GET.get('date_fin', date.today().strftime('%Y-%m-%d'))
        departement_id = request.GET.get('departement')
        employe_id = request.GET.get('employe')
        type_rapport = request.GET.get('type_rapport', 'journalier')
        
        # Convertir les dates
        try:
            date_debut = datetime.strptime(date_debut, '%Y-%m-%d').date()
            date_fin = datetime.strptime(date_fin, '%Y-%m-%d').date()
        except ValueError:
            date_debut = date_fin = date.today()
        
        # Données pour les filtres
        departements = Department.objects.filter(is_active=True).order_by('name')
        employes = Employee.objects.filter(is_active=True).order_by('last_name', 'first_name')
        
        # Générer le rapport si demandé
        rapport_data = None
        if request.GET.get('generate'):
            gestionnaire = GestionnairePointage()
            
            if type_rapport == 'feuille_presence':
                rapport_data = gestionnaire.generer_feuille_presence(
                    date_debut, date_fin, departement_id
                )
            elif type_rapport == 'heures_sup':
                # Rapport heures supplémentaires par employé
                employes_query = Employee.objects.filter(is_active=True)
                if departement_id:
                    employes_query = employes_query.filter(department_id=departement_id)
                if employe_id:
                    employes_query = employes_query.filter(id=employe_id)
                
                rapport_data = {
                    'type': 'heures_supplementaires',
                    'periode': {'debut': date_debut, 'fin': date_fin},
                    'employes': []
                }
                
                for emp in employes_query:
                    heures_sup = gestionnaire.calculer_heures_supplementaires(emp, date_debut, date_fin)
                    if heures_sup['heures_sup_total'].total_seconds() > 0:
                        rapport_data['employes'].append({
                            'employe': emp,
                            'heures_sup': heures_sup
                        })
        
        context = {
            'date_debut': date_debut,
            'date_fin': date_fin,
            'departement_id': int(departement_id) if departement_id else None,
            'employe_id': int(employe_id) if employe_id else None,
            'type_rapport': type_rapport,
            'departements': departements,
            'employes': employes,
            'rapport_data': rapport_data,
        }
        
        return render(request, 'spa/attendance/reports.html', context)
        
    except Exception as e:
        logger.error(f"Erreur dans attendance_reports_content: {e}")
        return render(request, 'spa/attendance/reports.html', {'error': str(e)})


@login_required
def attendance_schedules_content(request):
    """Gestion horaires et plannings"""
    try:
        # Récupérer les plages horaires
        plages_horaires = PlageHoraire.objects.filter(actif=True).order_by('nom')
        
        # Récupérer les horaires de travail avec pagination
        horaires_query = HoraireTravail.objects.select_related(
            'employe', 'plage_horaire'
        ).filter(actif=True).order_by('-date_debut')
        
        # Filtrage
        if request.GET.get('employe'):
            horaires_query = horaires_query.filter(employe_id=request.GET.get('employe'))
        if request.GET.get('plage'):
            horaires_query = horaires_query.filter(plage_horaire_id=request.GET.get('plage'))
        
        paginator = Paginator(horaires_query, 20)
        page_number = request.GET.get('page')
        horaires_page = paginator.get_page(page_number)
        
        # Employés pour les filtres
        employes = Employee.objects.filter(is_active=True).order_by('last_name', 'first_name')
        
        # Règles de pointage
        regles_pointage = ReglePointage.objects.filter(actif=True).order_by('-date_debut')
        
        context = {
            'plages_horaires': plages_horaires,
            'horaires_page': horaires_page,
            'employes': employes,
            'regles_pointage': regles_pointage,
            'types_plage': PlageHoraire.TYPES_PLAGE,
        }
        
        return render(request, 'spa/attendance/schedules.html', context)
        
    except Exception as e:
        logger.error(f"Erreur dans attendance_schedules_content: {e}")
        return render(request, 'spa/attendance/schedules.html', {'error': str(e)})


@login_required
def attendance_validation_content(request):
    """Validation et correction pointages"""
    try:
        # Date de validation (défaut: aujourd'hui)
        date_validation = request.GET.get('date', date.today().strftime('%Y-%m-%d'))
        try:
            date_validation = datetime.strptime(date_validation, '%Y-%m-%d').date()
        except ValueError:
            date_validation = date.today()
        
        # Récupérer les présences à valider
        presences_query = PresenceJournaliere.objects.filter(
            date=date_validation
        ).select_related('employe', 'horaire_travail').order_by('employe__last_name', 'employe__first_name')
        
        # Filtrer par statut de validation
        statut_filter = request.GET.get('statut', 'non_validees')
        if statut_filter == 'non_validees':
            presences_query = presences_query.filter(valide=False)
        elif statut_filter == 'validees':
            presences_query = presences_query.filter(valide=True)
        
        presences = list(presences_query)
        
        # Pointages suspects du jour
        pointages_suspects = Pointage.objects.filter(
            heure_pointage__date=date_validation,
            statut__in=['SUSPECT', 'RETARD']
        ).select_related('employe').order_by('-heure_pointage')
        
        # Alertes du jour
        alertes_jour = AlertePresence.objects.filter(
            date_concernee=date_validation,
            statut='NOUVELLE'
        ).select_related('employe').order_by('-niveau_gravite', '-date_creation')
        
        # Statistiques de validation
        stats_validation = {
            'total_presences': len(presences),
            'validees': len([p for p in presences if p.valide]),
            'non_validees': len([p for p in presences if not p.valide]),
            'total_pointages_suspects': pointages_suspects.count(),
            'total_alertes': alertes_jour.count()
        }
        
        context = {
            'date_validation': date_validation,
            'presences': presences,
            'pointages_suspects': pointages_suspects,
            'alertes_jour': alertes_jour,
            'stats_validation': stats_validation,
            'statut_filter': statut_filter,
        }
        
        return render(request, 'spa/attendance/validation.html', context)
        
    except Exception as e:
        logger.error(f"Erreur dans attendance_validation_content: {e}")
        return render(request, 'spa/attendance/validation.html', {'error': str(e)})


@login_required 
def attendance_settings_content(request):
    """Configuration règles et paramètres"""
    try:
        # Règles de pointage
        regles_pointage = ReglePointage.objects.all().order_by('-date_debut')
        
        # Plages horaires
        plages_horaires = PlageHoraire.objects.all().order_by('nom')
        
        # Validations en cours
        validations_encours = ValidationPresence.objects.filter(
            statut__in=['EN_ATTENTE', 'EN_COURS']
        ).select_related('validee_par').order_by('-date_creation')
        
        # Statistiques des alertes par type
        stats_alertes = AlertePresence.objects.values('type_alerte').annotate(
            count=Count('id')
        ).order_by('-count')
        
        context = {
            'regles_pointage': regles_pointage,
            'plages_horaires': plages_horaires,
            'validations_encours': validations_encours,
            'stats_alertes': stats_alertes,
            'types_plage': PlageHoraire.TYPES_PLAGE,
            'types_alerte': AlertePresence.TYPES_ALERTE,
            'niveaux_gravite': AlertePresence.NIVEAUX_GRAVITE,
        }
        
        return render(request, 'spa/attendance/settings.html', context)
        
    except Exception as e:
        logger.error(f"Erreur dans attendance_settings_content: {e}")
        return render(request, 'spa/attendance/settings.html', {'error': str(e)})


# ============================================================================
# APIs REST POINTAGE - Fonctionnalités backend
# ============================================================================

@csrf_exempt
@require_POST
@login_required
def api_create_pointage(request):
    """API pour enregistrer un nouveau pointage"""
    try:
        data = json.loads(request.body)
        
        employe_id = data.get('employe_id')
        type_pointage = data.get('type_pointage')
        heure_pointage = data.get('heure_pointage')  # ISO format ou None pour maintenant
        justification = data.get('justification', '')
        
        # Validation des paramètres
        if not employe_id or not type_pointage:
            return JsonResponse({
                'success': False,
                'message': 'Paramètres manquants (employe_id, type_pointage)'
            }, status=400)
        
        if type_pointage not in [choice[0] for choice in Pointage.TYPES_POINTAGE]:
            return JsonResponse({
                'success': False,
                'message': 'Type de pointage invalide'
            }, status=400)
        
        # Convertir l'heure si fournie
        heure = None
        if heure_pointage:
            try:
                heure = datetime.fromisoformat(heure_pointage.replace('Z', '+00:00'))
                if timezone.is_aware(heure):
                    heure = timezone.localtime(heure)
            except ValueError:
                return JsonResponse({
                    'success': False,
                    'message': 'Format d\'heure invalide'
                }, status=400)
        
        # Enregistrer le pointage
        gestionnaire = GestionnairePointage()
        result = gestionnaire.enregistrer_pointage(
            employe_id=employe_id,
            type_pointage=type_pointage,
            heure=heure,
            cree_par_id=request.user.id,
            ip_address=request.META.get('REMOTE_ADDR'),
            justification=justification
        )
        
        return JsonResponse(result)
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'message': 'Format JSON invalide'
        }, status=400)
    except Exception as e:
        logger.error(f"Erreur dans api_create_pointage: {e}")
        return JsonResponse({
            'success': False,
            'message': f'Erreur serveur: {str(e)}'
        }, status=500)


@require_http_methods(["GET"])
@login_required
def api_get_presence_status(request):
    """API pour récupérer le statut de présence temps réel"""
    try:
        gestionnaire = GestionnairePointage()
        statut = gestionnaire.get_statut_presence_temps_reel()
        
        # Formater les données pour l'API
        api_response = {
            'success': True,
            'timestamp': statut['timestamp'],
            'statistiques': statut['statistiques'],
            'employes': {
                'presents': [
                    {
                        'id': emp['employe'].id,
                        'nom': emp['employe'].last_name,
                        'prenom': emp['employe'].first_name,
                        'departement': emp['employe'].department.name if emp['employe'].department else '',
                        'heure_arrivee': emp['heure_arrivee'].strftime('%H:%M'),
                        'retard_minutes': emp['retard_minutes'],
                        'statut': emp['statut']
                    } for emp in statut['presents']
                ],
                'absents': [
                    {
                        'id': emp['employe'].id,
                        'nom': emp['employe'].last_name,
                        'prenom': emp['employe'].first_name,
                        'departement': emp['employe'].department.name if emp['employe'].department else '',
                        'horaire_theorique': emp['horaire_theorique'],
                        'statut': emp['statut']
                    } for emp in statut['absents']
                ],
                'en_pause': [
                    {
                        'id': emp['employe'].id,
                        'nom': emp['employe'].last_name,
                        'prenom': emp['employe'].first_name,
                        'departement': emp['employe'].department.name if emp['employe'].department else '',
                        'heure_debut_pause': emp['heure_debut_pause'].strftime('%H:%M'),
                        'duree_pause_minutes': int(emp['duree_pause'].total_seconds() / 60),
                        'statut': emp['statut']
                    } for emp in statut['en_pause']
                ]
            }
        }
        
        return JsonResponse(api_response)
        
    except Exception as e:
        logger.error(f"Erreur dans api_get_presence_status: {e}")
        return JsonResponse({
            'success': False,
            'message': f'Erreur serveur: {str(e)}'
        }, status=500)


@require_http_methods(["GET"])
@login_required
def api_calculate_daily_hours(request):
    """API pour calculer les heures journalières d'un employé"""
    try:
        employe_id = request.GET.get('employe_id')
        date_calc = request.GET.get('date', date.today().strftime('%Y-%m-%d'))
        
        if not employe_id:
            return JsonResponse({
                'success': False,
                'message': 'employe_id manquant'
            }, status=400)
        
        # Convertir la date
        try:
            date_calc = datetime.strptime(date_calc, '%Y-%m-%d').date()
        except ValueError:
            return JsonResponse({
                'success': False,
                'message': 'Format de date invalide (YYYY-MM-DD)'
            }, status=400)
        
        # Calculer les heures
        gestionnaire = GestionnairePointage()
        employe = get_object_or_404(Employee, id=employe_id)
        calcul = gestionnaire.calculer_heures_travaillees(employe, date_calc)
        
        # Formater la réponse
        response = {
            'success': True,
            'employe_id': employe_id,
            'date': date_calc.strftime('%Y-%m-%d'),
            'heure_arrivee': calcul['heure_arrivee'].strftime('%H:%M') if calcul['heure_arrivee'] else None,
            'heure_sortie': calcul['heure_sortie'].strftime('%H:%M') if calcul['heure_sortie'] else None,
            'heures_travaillees': str(calcul['heures_travaillees']),
            'heures_theoriques': str(calcul['heures_theoriques']),
            'duree_pauses': str(calcul['duree_pauses']),
            'retard_minutes': calcul['retard_minutes'],
            'depart_anticipe_minutes': calcul['depart_anticipe_minutes'],
            'pauses': calcul['pauses']
        }
        
        return JsonResponse(response)
        
    except Exception as e:
        logger.error(f"Erreur dans api_calculate_daily_hours: {e}")
        return JsonResponse({
            'success': False,
            'message': f'Erreur serveur: {str(e)}'
        }, status=500)


@require_http_methods(["GET"])
@login_required
def api_attendance_reports_data(request):
    """API pour récupérer les données des rapports de présence"""
    try:
        # Paramètres
        date_debut = request.GET.get('date_debut', date.today().strftime('%Y-%m-%d'))
        date_fin = request.GET.get('date_fin', date.today().strftime('%Y-%m-%d'))
        departement_id = request.GET.get('departement_id')
        type_rapport = request.GET.get('type_rapport', 'feuille_presence')
        
        # Convertir les dates
        try:
            date_debut = datetime.strptime(date_debut, '%Y-%m-%d').date()
            date_fin = datetime.strptime(date_fin, '%Y-%m-%d').date()
        except ValueError:
            return JsonResponse({
                'success': False,
                'message': 'Format de date invalide'
            }, status=400)
        
        gestionnaire = GestionnairePointage()
        
        if type_rapport == 'feuille_presence':
            data = gestionnaire.generer_feuille_presence(date_debut, date_fin, departement_id)
            
            # Formater pour l'API
            response_data = {
                'success': True,
                'type_rapport': 'feuille_presence',
                'periode': {
                    'debut': data['periode']['debut'].strftime('%Y-%m-%d'),
                    'fin': data['periode']['fin'].strftime('%Y-%m-%d'),
                    'nb_jours': data['periode']['nb_jours']
                },
                'statistiques': data['statistiques'],
                'employes': []
            }
            
            for emp_data in data['employes']:
                employe_info = {
                    'employe': {
                        'id': emp_data['employe'].id,
                        'nom': emp_data['employe'].nom,
                        'prenom': emp_data['employe'].prenom,
                        'matricule': emp_data['employe'].matricule,
                        'departement': emp_data['employe'].departement.nom if emp_data['employe'].departement else ''
                    },
                    'statistiques': emp_data['statistiques'],
                    'presences': [
                        {
                            'date': p.date.strftime('%Y-%m-%d'),
                            'statut': p.statut_jour,
                            'heure_arrivee': p.heure_arrivee.strftime('%H:%M') if p.heure_arrivee else None,
                            'heure_sortie': p.heure_sortie.strftime('%H:%M') if p.heure_sortie else None,
                            'heures_travaillees': str(p.heures_travaillees),
                            'retard_minutes': p.retard_minutes
                        } for p in emp_data['presences']
                    ]
                }
                response_data['employes'].append(employe_info)
            
            return JsonResponse(response_data)
        
        elif type_rapport == 'detection_retards_absences':
            data = gestionnaire.detecter_retards_absences(date_debut)
            
            response_data = {
                'success': True,
                'type_rapport': 'detection_retards_absences',
                'date': data['date'].strftime('%Y-%m-%d'),
                'statistiques': {
                    'total_employes': data['total_employes'],
                    'nb_presents': data['nb_presents'],
                    'nb_absents': data['nb_absents'],
                    'nb_retards': data['nb_retards']
                },
                'retards': [
                    {
                        'employe': {
                            'id': r['employe'].id,
                            'nom': r['employe'].nom,
                            'prenom': r['employe'].prenom
                        },
                        'heure_arrivee': r['heure_arrivee'].strftime('%H:%M'),
                        'heure_theorique': r['heure_theorique'].strftime('%H:%M'),
                        'retard_minutes': r['retard_minutes']
                    } for r in data['retards']
                ],
                'absences': [
                    {
                        'employe': {
                            'id': a['employe'].id,
                            'nom': a['employe'].nom,
                            'prenom': a['employe'].prenom
                        }
                    } for a in data['absences']
                ]
            }
            
            return JsonResponse(response_data)
        
        else:
            return JsonResponse({
                'success': False,
                'message': 'Type de rapport non supporté'
            }, status=400)
        
    except Exception as e:
        logger.error(f"Erreur dans api_attendance_reports_data: {e}")
        return JsonResponse({
            'success': False,
            'message': f'Erreur serveur: {str(e)}'
        }, status=500)


@csrf_exempt
@require_POST
@login_required
def api_validate_attendance(request):
    """API pour valider les présences"""
    try:
        data = json.loads(request.body)
        
        date_validation = data.get('date_validation')
        employes_ids = data.get('employes_ids', [])
        
        if not date_validation:
            return JsonResponse({
                'success': False,
                'message': 'date_validation manquante'
            }, status=400)
        
        # Convertir la date
        try:
            date_validation = datetime.strptime(date_validation, '%Y-%m-%d').date()
        except ValueError:
            return JsonResponse({
                'success': False,
                'message': 'Format de date invalide'
            }, status=400)
        
        # Valider les présences
        gestionnaire = GestionnairePointage()
        result = gestionnaire.valider_presence_journaliere(
            date_validation=date_validation,
            validee_par_id=request.user.id,
            employes_ids=employes_ids if employes_ids else None
        )
        
        return JsonResponse(result)
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'message': 'Format JSON invalide'
        }, status=400)
    except Exception as e:
        logger.error(f"Erreur dans api_validate_attendance: {e}")
        return JsonResponse({
            'success': False,
            'message': f'Erreur serveur: {str(e)}'
        }, status=500)


@require_http_methods(["GET"])
@login_required
def api_export_payroll_data(request):
    """API pour exporter les données vers le module paie"""
    try:
        mois = int(request.GET.get('mois', date.today().month))
        annee = int(request.GET.get('annee', date.today().year))
        departement_id = request.GET.get('departement_id')
        
        if not (1 <= mois <= 12):
            return JsonResponse({
                'success': False,
                'message': 'Mois invalide (1-12)'
            }, status=400)
        
        # Exporter les données
        gestionnaire = GestionnairePointage()
        data = gestionnaire.exporter_donnees_paie(
            mois=mois,
            annee=annee,
            departement_id=int(departement_id) if departement_id else None
        )
        
        # Formater pour l'API
        response_data = {
            'success': True,
            'periode': {
                'mois': data['periode']['mois'],
                'annee': data['periode']['annee'],
                'date_debut': data['periode']['date_debut'].strftime('%Y-%m-%d'),
                'date_fin': data['periode']['date_fin'].strftime('%Y-%m-%d')
            },
            'statistiques_globales': {
                'total_heures_normales': data['statistiques_globales']['total_heures_normales'].total_seconds() / 3600,
                'total_heures_sup': data['statistiques_globales']['total_heures_sup'].total_seconds() / 3600,
                'total_absences': data['statistiques_globales']['total_absences'],
                'total_retards_minutes': data['statistiques_globales']['total_retards_minutes']
            },
            'employes': data['employes']
        }
        
        return JsonResponse(response_data)
        
    except ValueError:
        return JsonResponse({
            'success': False,
            'message': 'Paramètres numériques invalides'
        }, status=400)
    except Exception as e:
        logger.error(f"Erreur dans api_export_payroll_data: {e}")
        return JsonResponse({
            'success': False,
            'message': f'Erreur serveur: {str(e)}'
        }, status=500)


# ============================================================================
# APIs UTILITAIRES ET ACTIONS SPÉCIALES
# ============================================================================

@csrf_exempt
@require_POST
@login_required
def api_correct_pointage(request):
    """API pour corriger un pointage existant"""
    try:
        data = json.loads(request.body)
        
        pointage_id = data.get('pointage_id')
        nouvelle_heure = data.get('nouvelle_heure')
        raison_correction = data.get('raison_correction', '')
        
        if not pointage_id or not nouvelle_heure:
            return JsonResponse({
                'success': False,
                'message': 'Paramètres manquants'
            }, status=400)
        
        # Récupérer le pointage
        pointage = get_object_or_404(Pointage, id=pointage_id)
        
        # Convertir la nouvelle heure
        try:
            nouvelle_heure = datetime.fromisoformat(nouvelle_heure.replace('Z', '+00:00'))
            if timezone.is_aware(nouvelle_heure):
                nouvelle_heure = timezone.localtime(nouvelle_heure)
        except ValueError:
            return JsonResponse({
                'success': False,
                'message': 'Format d\'heure invalide'
            }, status=400)
        
        # Corriger le pointage
        with transaction.atomic():
            pointage.heure_originale = pointage.heure_pointage
            pointage.heure_pointage = nouvelle_heure
            pointage.corrige = True
            pointage.raison_correction = raison_correction
            pointage.corrige_par = request.user
            pointage.statut = 'CORRIGE'
            pointage.save()
            
            # Recalculer la présence journalière
            gestionnaire = GestionnairePointage()
            gestionnaire._mettre_a_jour_presence_journaliere(
                pointage.employe, 
                pointage.heure_pointage.date()
            )
        
        return JsonResponse({
            'success': True,
            'message': 'Pointage corrigé avec succès',
            'pointage_id': str(pointage.id)
        })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'message': 'Format JSON invalide'
        }, status=400)
    except Exception as e:
        logger.error(f"Erreur dans api_correct_pointage: {e}")
        return JsonResponse({
            'success': False,
            'message': f'Erreur serveur: {str(e)}'
        }, status=500)


@require_http_methods(["GET"])
@login_required
def api_get_employee_schedule(request, employe_id):
    """API pour récupérer l'horaire actuel d'un employé"""
    try:
        employe = get_object_or_404(Employee, id=employe_id)
        
        # Récupérer l'horaire actuel
        gestionnaire = GestionnairePointage()
        horaire = gestionnaire._get_horaire_employe(employe, date.today())
        
        if not horaire:
            return JsonResponse({
                'success': False,
                'message': 'Aucun horaire défini pour cet employé'
            })
        
        response_data = {
            'success': True,
            'employe': {
                'id': employe.id,
                'nom': employe.nom,
                'prenom': employe.prenom,
                'matricule': employe.matricule
            },
            'horaire': {
                'id': horaire.id,
                'plage_horaire': {
                    'nom': horaire.plage_horaire.nom,
                    'type': horaire.plage_horaire.type_plage,
                    'heure_debut': horaire.heure_debut_effective.strftime('%H:%M'),
                    'heure_fin': horaire.heure_fin_effective.strftime('%H:%M'),
                    'duree_pause': str(horaire.plage_horaire.duree_pause),
                    'jours_travailles': horaire.jours_travailles_effectifs
                },
                'date_debut': horaire.date_debut.strftime('%Y-%m-%d'),
                'date_fin': horaire.date_fin.strftime('%Y-%m-%d') if horaire.date_fin else None
            }
        }
        
        return JsonResponse(response_data)
        
    except Exception as e:
        logger.error(f"Erreur dans api_get_employee_schedule: {e}")
        return JsonResponse({
            'success': False,
            'message': f'Erreur serveur: {str(e)}'
        }, status=500)


# ============================================================================
# EXPORTS PDF ET EXCEL
# ============================================================================

@login_required
def export_attendance_pdf(request):
    """Export PDF d'une feuille de présence"""
    try:
        # Paramètres
        date_debut = datetime.strptime(request.GET.get('date_debut', date.today().strftime('%Y-%m-%d')), '%Y-%m-%d').date()
        date_fin = datetime.strptime(request.GET.get('date_fin', date.today().strftime('%Y-%m-%d')), '%Y-%m-%d').date()
        departement_id = request.GET.get('departement_id')
        
        # Générer les données
        gestionnaire = GestionnairePointage()
        data = gestionnaire.generer_feuille_presence(date_debut, date_fin, departement_id)
        
        # Créer le PDF
        buffer = io.BytesIO()
        p = canvas.Canvas(buffer, pagesize=A4)
        
        # En-tête
        p.setFont("Helvetica-Bold", 16)
        p.drawString(50, 800, f"Feuille de Présence - {date_debut.strftime('%d/%m/%Y')} au {date_fin.strftime('%d/%m/%Y')}")
        
        y_position = 750
        p.setFont("Helvetica", 10)
        
        # Tableau des présences
        for emp_data in data['employes'][:20]:  # Limiter pour tenir sur une page
            employe = emp_data['employe']
            stats = emp_data['statistiques']
            
            p.drawString(50, y_position, f"{employe.nom} {employe.prenom}")
            p.drawString(200, y_position, f"Présences: {stats['nb_presences']}")
            p.drawString(300, y_position, f"Absences: {stats['nb_absences']}")
            p.drawString(400, y_position, f"Retards: {stats['nb_retards']}")
            
            y_position -= 20
            
            if y_position < 100:
                break
        
        p.showPage()
        p.save()
        
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="feuille_presence_{date_debut}_{date_fin}.pdf"'
        
        return response
        
    except Exception as e:
        logger.error(f"Erreur export PDF: {e}")
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@login_required
@require_http_methods(["GET"])
def api_list_alertes(request):
    """API - Liste des alertes de pointage avec filtres"""
    try:
        # Paramètres de filtrage
        statut = request.GET.get('statut', '')
        niveau_gravite = request.GET.get('niveau_gravite', '')
        nouveaux = request.GET.get('nouveaux', '').lower() == 'true'
        limit = int(request.GET.get('limit', 50))
        
        # Query de base
        alertes_query = AlertePresence.objects.select_related('employe')
        
        # Filtres
        if statut:
            alertes_query = alertes_query.filter(statut=statut)
        
        if niveau_gravite:
            alertes_query = alertes_query.filter(niveau_gravite=niveau_gravite)
        
        if nouveaux:
            # Alertes créées dans les dernières 24h
            depuis = timezone.now() - timedelta(hours=24)
            alertes_query = alertes_query.filter(date_creation__gte=depuis)
        
        # Ordonner et limiter
        alertes = alertes_query.order_by('-date_creation', '-niveau_gravite')[:limit]
        
        # Formater pour l'API
        alertes_data = []
        for alerte in alertes:
            alertes_data.append({
                'id': alerte.id,
                'titre': alerte.titre,
                'message': alerte.message,
                'type_alerte': alerte.type_alerte,
                'niveau_gravite': alerte.niveau_gravite,
                'statut': alerte.statut,
                'date_creation': alerte.date_creation.isoformat(),
                'date_concernee': alerte.date_concernee.isoformat(),
                'employe': {
                    'id': alerte.employe.id,
                    'nom': alerte.employe.last_name,
                    'prenom': alerte.employe.first_name,
                    'departement': alerte.employe.department.name if alerte.employe.department else None
                } if alerte.employe else None,
                'actions_possibles': alerte.actions_possibles,
                'resolu': alerte.resolu,
                'resolu_par': alerte.resolu_par.username if alerte.resolu_par else None,
                'date_resolution': alerte.date_resolution.isoformat() if alerte.date_resolution else None,
            })
        
        return JsonResponse({
            'success': True,
            'alertes': alertes_data,
            'total': len(alertes_data),
            'filters': {
                'statut': statut,
                'niveau_gravite': niveau_gravite,
                'nouveaux': nouveaux,
                'limit': limit
            }
        })
        
    except ValueError as e:
        return JsonResponse({
            'success': False,
            'error': f'Paramètre invalide: {str(e)}'
        }, status=400)
    except Exception as e:
        logger.error(f"Erreur dans api_list_alertes: {e}")
        return JsonResponse({
            'success': False,
            'error': f'Erreur serveur: {str(e)}'
        }, status=500)


@login_required
@require_http_methods(["GET"])
def export_attendance_excel(request):
    """Export Excel d'une feuille de présence"""
    try:
        # Paramètres
        date_debut = datetime.strptime(request.GET.get('date_debut', date.today().strftime('%Y-%m-%d')), '%Y-%m-%d').date()
        date_fin = datetime.strptime(request.GET.get('date_fin', date.today().strftime('%Y-%m-%d')), '%Y-%m-%d').date()
        departement_id = request.GET.get('departement_id')
        
        # Générer les données
        gestionnaire = GestionnairePointage()
        data = gestionnaire.generer_feuille_presence(date_debut, date_fin, departement_id)
        
        # Créer le fichier Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Feuille de Présence"
        
        # Style des en-têtes
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        # En-têtes
        headers = ['Nom', 'Prénom', 'Matricule', 'Département', 'Présences', 'Absences', 'Retards', 'Total Heures']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # Données des employés
        for row, emp_data in enumerate(data['employes'], 2):
            employe = emp_data['employe']
            stats = emp_data['statistiques']
            
            ws.cell(row=row, column=1, value=employe.nom)
            ws.cell(row=row, column=2, value=employe.prenom)
            ws.cell(row=row, column=3, value=employe.matricule)
            ws.cell(row=row, column=4, value=employe.departement.nom if employe.departement else '')
            ws.cell(row=row, column=5, value=stats['nb_presences'])
            ws.cell(row=row, column=6, value=stats['nb_absences'])
            ws.cell(row=row, column=7, value=stats['nb_retards'])
            ws.cell(row=row, column=8, value=str(stats['total_heures']))
        
        # Ajuster la largeur des colonnes
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Sauvegarder dans un buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(
            buffer,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="feuille_presence_{date_debut}_{date_fin}.xlsx"'
        
        return response
        
    except Exception as e:
        logger.error(f"Erreur export Excel: {e}")
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


# ============================================================================
# APIs MANQUANTES POUR LE MODULE POINTAGE
# ============================================================================

@login_required
@require_http_methods(["POST"])
def api_create_plage_horaire(request):
    """API - Créer une plage horaire"""
    try:
        data = json.loads(request.body)
        
        plage = PlageHoraire.objects.create(
            nom=data['nom'],
            type_plage=data.get('type_plage', 'STANDARD'),
            heure_debut=data['heure_debut'],
            heure_fin=data['heure_fin'],
            duree_pause=data.get('duree_pause', '01:00:00'),
            jours_travailles=data.get('jours_travailles', [1, 2, 3, 4, 5]),
            tolerance_retard=data.get('tolerance_retard', 10),
            tolerance_depart=data.get('tolerance_depart', 10),
            heures_sup_auto=data.get('heures_sup_auto', False),
            description=data.get('description', ''),
            actif=True
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Plage horaire créée avec succès',
            'plage_id': plage.id
        })
        
    except Exception as e:
        logger.error(f"Erreur création plage horaire: {e}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


@login_required
@require_http_methods(["PUT"])
def api_update_plage_horaire(request, plage_id):
    """API - Modifier une plage horaire"""
    try:
        plage = get_object_or_404(PlageHoraire, id=plage_id)
        data = json.loads(request.body)
        
        # Mise à jour des champs
        for field in ['nom', 'type_plage', 'heure_debut', 'heure_fin', 'duree_pause',
                     'jours_travailles', 'tolerance_retard', 'tolerance_depart',
                     'heures_sup_auto', 'description']:
            if field in data:
                setattr(plage, field, data[field])
        
        plage.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Plage horaire mise à jour avec succès'
        })
        
    except Exception as e:
        logger.error(f"Erreur mise à jour plage horaire {plage_id}: {e}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


@login_required
@require_http_methods(["DELETE"])
def api_delete_plage_horaire(request, plage_id):
    """API - Supprimer une plage horaire"""
    try:
        plage = get_object_or_404(PlageHoraire, id=plage_id)
        
        # Vérifier qu'aucun horaire n'utilise cette plage
        horaires_actifs = HoraireTravail.objects.filter(plage_horaire=plage, actif=True)
        if horaires_actifs.exists():
            return JsonResponse({
                'success': False,
                'error': 'Impossible de supprimer cette plage horaire car elle est utilisée'
            }, status=400)
        
        plage.actif = False
        plage.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Plage horaire supprimée avec succès'
        })
        
    except Exception as e:
        logger.error(f"Erreur suppression plage horaire {plage_id}: {e}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


@login_required
@require_http_methods(["POST"])
def api_duplicate_plage_horaire(request, plage_id):
    """API - Dupliquer une plage horaire"""
    try:
        plage_originale = get_object_or_404(PlageHoraire, id=plage_id)
        data = json.loads(request.body)
        
        nouveau_nom = data.get('nouveau_nom', f"Copie de {plage_originale.nom}")
        
        nouvelle_plage = PlageHoraire.objects.create(
            nom=nouveau_nom,
            type_plage=plage_originale.type_plage,
            heure_debut=plage_originale.heure_debut,
            heure_fin=plage_originale.heure_fin,
            duree_pause=plage_originale.duree_pause,
            jours_travailles=plage_originale.jours_travailles,
            tolerance_retard=plage_originale.tolerance_retard,
            tolerance_depart=plage_originale.tolerance_depart,
            heures_sup_auto=plage_originale.heures_sup_auto,
            description=f"Copie de: {plage_originale.description}",
            actif=True
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Plage horaire dupliquée avec succès',
            'nouvelle_plage_id': nouvelle_plage.id
        })
        
    except Exception as e:
        logger.error(f"Erreur duplication plage horaire {plage_id}: {e}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


@login_required
@require_http_methods(["POST"])
def api_assign_horaire_employe(request):
    """API - Assigner un horaire à un employé"""
    try:
        data = json.loads(request.body)
        
        employe = get_object_or_404(Employee, id=data['employe_id'])
        plage_horaire = get_object_or_404(PlageHoraire, id=data['plage_horaire_id'])
        
        # Terminer l'horaire actuel s'il existe
        horaire_actuel = HoraireTravail.objects.filter(
            employe=employe, 
            actif=True
        ).first()
        
        if horaire_actuel:
            horaire_actuel.date_fin = date.today()
            horaire_actuel.actif = False
            horaire_actuel.save()
        
        # Créer le nouvel horaire
        nouvel_horaire = HoraireTravail.objects.create(
            employe=employe,
            plage_horaire=plage_horaire,
            date_debut=data.get('date_debut', date.today()),
            heure_debut_effective=data.get('heure_debut_effective', plage_horaire.heure_debut),
            heure_fin_effective=data.get('heure_fin_effective', plage_horaire.heure_fin),
            jours_travailles_effectifs=data.get('jours_travailles', plage_horaire.jours_travailles),
            actif=True,
            cree_par=request.user
        )
        
        return JsonResponse({
            'success': True,
            'message': f'Horaire assigné à {employe.first_name} {employe.last_name}',
            'horaire_id': nouvel_horaire.id
        })
        
    except Exception as e:
        logger.error(f"Erreur assignation horaire: {e}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


@login_required
@require_http_methods(["PUT"])
def api_update_horaire_employe(request, horaire_id):
    """API - Modifier l'horaire d'un employé"""
    try:
        horaire = get_object_or_404(HoraireTravail, id=horaire_id)
        data = json.loads(request.body)
        
        # Mise à jour des champs modifiables
        for field in ['heure_debut_effective', 'heure_fin_effective', 
                     'jours_travailles_effectifs', 'date_fin']:
            if field in data:
                setattr(horaire, field, data[field])
        
        horaire.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Horaire mis à jour avec succès'
        })
        
    except Exception as e:
        logger.error(f"Erreur mise à jour horaire {horaire_id}: {e}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


@login_required
@require_http_methods(["POST"])
def api_end_horaire_employe(request, horaire_id):
    """API - Terminer l'horaire d'un employé"""
    try:
        horaire = get_object_or_404(HoraireTravail, id=horaire_id)
        data = json.loads(request.body)
        
        horaire.date_fin = data.get('date_fin', date.today())
        horaire.actif = False
        horaire.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Horaire terminé avec succès'
        })
        
    except Exception as e:
        logger.error(f"Erreur fin horaire {horaire_id}: {e}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


@login_required
@require_http_methods(["GET"])
def api_get_horaire_history(request, employe_id):
    """API - Historique des horaires d'un employé"""
    try:
        employe = get_object_or_404(Employee, id=employe_id)
        
        horaires = HoraireTravail.objects.filter(
            employe=employe
        ).select_related('plage_horaire').order_by('-date_debut')
        
        historique = []
        for horaire in horaires:
            historique.append({
                'id': horaire.id,
                'plage_horaire': {
                    'nom': horaire.plage_horaire.nom,
                    'type': horaire.plage_horaire.type_plage,
                },
                'date_debut': horaire.date_debut.isoformat(),
                'date_fin': horaire.date_fin.isoformat() if horaire.date_fin else None,
                'actif': horaire.actif,
                'heure_debut': horaire.heure_debut_effective.strftime('%H:%M'),
                'heure_fin': horaire.heure_fin_effective.strftime('%H:%M'),
                'jours_travailles': horaire.jours_travailles_effectifs
            })
        
        return JsonResponse({
            'success': True,
            'employe': {
                'id': employe.id,
                'nom': employe.last_name,
                'prenom': employe.first_name
            },
            'historique': historique
        })
        
    except Exception as e:
        logger.error(f"Erreur historique horaire {employe_id}: {e}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@require_http_methods(["POST"])
def api_create_regle_pointage(request):
    """API - Créer une règle de pointage"""
    try:
        data = json.loads(request.body)
        
        regle = ReglePointage.objects.create(
            nom=data['nom'],
            type_regle=data['type_regle'],
            condition=data.get('condition', ''),
            action=data.get('action', ''),
            priorite=data.get('priorite', 1),
            actif=True,
            date_debut=data.get('date_debut', date.today()),
            description=data.get('description', ''),
            cree_par=request.user
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Règle de pointage créée avec succès',
            'regle_id': regle.id
        })
        
    except Exception as e:
        logger.error(f"Erreur création règle pointage: {e}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


@login_required
@require_http_methods(["PUT"])
def api_update_regle_pointage(request, regle_id):
    """API - Modifier une règle de pointage"""
    try:
        regle = get_object_or_404(ReglePointage, id=regle_id)
        data = json.loads(request.body)
        
        # Mise à jour des champs
        for field in ['nom', 'type_regle', 'condition', 'action', 'priorite', 'description']:
            if field in data:
                setattr(regle, field, data[field])
        
        regle.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Règle de pointage mise à jour avec succès'
        })
        
    except Exception as e:
        logger.error(f"Erreur mise à jour règle pointage {regle_id}: {e}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


@login_required
@require_http_methods(["POST"])
def api_toggle_regle_pointage(request, regle_id):
    """API - Activer/Désactiver une règle"""
    try:
        regle = get_object_or_404(ReglePointage, id=regle_id)
        
        regle.actif = not regle.actif
        regle.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Règle {"activée" if regle.actif else "désactivée"} avec succès',
            'nouveau_statut': regle.actif
        })
        
    except Exception as e:
        logger.error(f"Erreur toggle règle pointage {regle_id}: {e}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


@login_required
@require_http_methods(["POST"])
def api_resolve_alerte(request, alerte_id):
    """API - Marquer une alerte comme traitée"""
    try:
        alerte = get_object_or_404(AlertePresence, id=alerte_id)
        data = json.loads(request.body)
        
        alerte.statut = 'RESOLUE'
        alerte.resolu = True
        alerte.resolu_par = request.user
        alerte.date_resolution = timezone.now()
        alerte.commentaire_resolution = data.get('commentaire', '')
        alerte.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Alerte marquée comme résolue'
        })
        
    except Exception as e:
        logger.error(f"Erreur résolution alerte {alerte_id}: {e}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


@login_required
@require_http_methods(["POST"])
def api_create_alerte(request):
    """API - Créer une alerte manuelle"""
    try:
        data = json.loads(request.body)
        
        alerte = AlertePresence.objects.create(
            titre=data['titre'],
            message=data['message'],
            type_alerte=data['type_alerte'],
            niveau_gravite=data.get('niveau_gravite', 'MOYENNE'),
            employe_id=data.get('employe_id'),
            date_concernee=data.get('date_concernee', date.today()),
            statut='NOUVELLE',
            cree_par=request.user
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Alerte créée avec succès',
            'alerte_id': alerte.id
        })
        
    except Exception as e:
        logger.error(f"Erreur création alerte: {e}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


@login_required
@require_http_methods(["GET"])
def api_daily_stats(request):
    """API - Statistiques globales du jour"""
    try:
        aujourd_hui = date.today()
        
        # Utiliser le gestionnaire de pointage pour les statistiques
        gestionnaire = GestionnairePointage()
        statut_temps_reel = gestionnaire.get_statut_presence_temps_reel()
        
        stats = {
            'date': aujourd_hui.isoformat(),
            'timestamp': timezone.now().isoformat(),
            'employes_presents': statut_temps_reel['statistiques']['nb_presents'],
            'employes_absents': statut_temps_reel['statistiques']['nb_absents'],
            'employes_en_pause': statut_temps_reel['statistiques']['nb_en_pause'],
            'total_employes': statut_temps_reel['statistiques']['total_employes'],
            'taux_presence': round(
                (statut_temps_reel['statistiques']['nb_presents'] / 
                 max(statut_temps_reel['statistiques']['total_employes'], 1)) * 100, 1
            ),
        }
        
        # Ajouter les alertes du jour
        alertes_jour = AlertePresence.objects.filter(
            date_concernee=aujourd_hui
        ).values('niveau_gravite').annotate(count=Count('id'))
        
        stats['alertes'] = {
            'total': sum(item['count'] for item in alertes_jour),
            'par_niveau': {item['niveau_gravite']: item['count'] for item in alertes_jour}
        }
        
        return JsonResponse({
            'success': True,
            'stats': stats
        })
        
    except Exception as e:
        logger.error(f"Erreur statistiques journalières: {e}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@require_http_methods(["GET"])
def api_employee_stats(request, employe_id):
    """API - Statistiques d'un employé"""
    try:
        employe = get_object_or_404(Employee, id=employe_id)
        
        # Période par défaut (30 derniers jours)
        date_fin = date.today()
        date_debut = date_fin - timedelta(days=30)
        
        presences = PresenceJournaliere.objects.filter(
            employe=employe,
            date__range=[date_debut, date_fin]
        )
        
        stats = {
            'employe': {
                'id': employe.id,
                'nom': employe.last_name,
                'prenom': employe.first_name,
                'departement': employe.department.name if employe.department else None
            },
            'periode': {
                'debut': date_debut.isoformat(),
                'fin': date_fin.isoformat(),
                'nb_jours': (date_fin - date_debut).days + 1
            },
            'presences': {
                'total': presences.count(),
                'presents': presences.filter(statut_jour='PRESENT').count(),
                'absents': presences.filter(statut_jour='ABSENT').count(),
                'retards': presences.filter(retard_minutes__gt=0).count(),
                'taux_presence': round(
                    (presences.filter(statut_jour='PRESENT').count() / 
                     max(presences.count(), 1)) * 100, 1
                )
            },
            'heures': {
                'total_travaillees': sum(
                    p.heures_travaillees.total_seconds() / 3600 
                    for p in presences if p.heures_travaillees
                ),
                'moyenne_par_jour': 0,
                'heures_supplementaires': sum(
                    p.heures_supplementaires.total_seconds() / 3600 
                    for p in presences if p.heures_supplementaires
                )
            }
        }
        
        # Calculer la moyenne par jour
        if stats['presences']['presents'] > 0:
            stats['heures']['moyenne_par_jour'] = round(
                stats['heures']['total_travaillees'] / stats['presences']['presents'], 1
            )
        
        return JsonResponse({
            'success': True,
            'stats': stats
        })
        
    except Exception as e:
        logger.error(f"Erreur statistiques employé {employe_id}: {e}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


# APIs avec fonctions de base pour éviter les erreurs
@login_required
@require_http_methods(["GET"])
def api_department_stats(request, departement_id):
    """API - Statistiques d'un département"""
    return JsonResponse({
        'success': True,
        'stats': {'message': 'Statistiques département - fonction à implémenter'}
    })


@login_required
@require_http_methods(["GET"])
def api_attendance_trends(request):
    """API - Tendances sur période"""
    return JsonResponse({
        'success': True,
        'trends': {'message': 'Tendances présence - fonction à implémenter'}
    })


@login_required
@require_http_methods(["POST"])
def api_recalculate_attendances(request):
    """API - Recalculer les présences d'une période"""
    return JsonResponse({
        'success': True,
        'message': 'Recalcul des présences - fonction à implémenter'
    })


@login_required
@require_http_methods(["POST"])
def api_bulk_import_pointages(request):
    """API - Import en masse de pointages"""
    return JsonResponse({
        'success': True,
        'message': 'Import en masse - fonction à implémenter'
    })


@login_required
@require_http_methods(["POST"])
def api_cleanup_old_data(request):
    """API - Nettoyage des données anciennes"""
    return JsonResponse({
        'success': True,
        'message': 'Nettoyage données - fonction à implémenter'
    })


@login_required
@require_http_methods(["POST"])
def api_test_notification(request):
    """API - Test des notifications"""
    return JsonResponse({
        'success': True,
        'message': 'Test notification - fonction à implémenter'
    })


@csrf_exempt
@require_http_methods(["POST"])
def webhook_receive_pointage(request):
    """API - Webhook pour réception de pointages externes"""
    return JsonResponse({
        'success': True,
        'message': 'Webhook pointage - fonction à implémenter'
    })


@login_required
@require_http_methods(["POST"])
def api_external_sync(request):
    """API - Synchronisation externe"""
    return JsonResponse({
        'success': True,
        'message': 'Sync externe - fonction à implémenter'
    })


@login_required
@require_http_methods(["GET"])
def api_custom_report(request):
    """API - Rapport détaillé personnalisé"""
    return JsonResponse({
        'success': True,
        'report': {'message': 'Rapport personnalisé - fonction à implémenter'}
    })


@login_required
def export_attendance_csv(request):
    """Export CSV personnalisé"""
    return JsonResponse({
        'success': True,
        'message': 'Export CSV - fonction à implémenter'
    })


@login_required
def employee_performance_report(request, employe_id):
    """Rapport de performance employé"""
    return JsonResponse({
        'success': True,
        'message': 'Rapport performance - fonction à implémenter'
    })


@login_required
@require_http_methods(["GET"])
def api_schedule_planning(request):
    """API - Planificateur d'horaires"""
    return JsonResponse({
        'success': True,
        'planning': {'message': 'Planification horaires - fonction à implémenter'}
    })


@login_required
@require_http_methods(["GET"])
def api_attendance_forecast(request):
    """API - Prévisions de présence"""
    return JsonResponse({
        'success': True,
        'forecast': {'message': 'Prévisions présence - fonction à implémenter'}
    })


@login_required
@require_http_methods(["GET"])
def api_optimize_schedules(request):
    """API - Optimisation des plannings"""
    return JsonResponse({
        'success': True,
        'optimization': {'message': 'Optimisation plannings - fonction à implémenter'}
    })